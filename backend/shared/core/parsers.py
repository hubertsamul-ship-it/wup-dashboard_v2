"""
Shared parsers — wrappery nad istniejącym kodem ninja/mrpips01_parser.py
oraz nad logiką openpyxl z extract_data.py.

Funkcje eksportowane przez FastAPI endpoints:
  parse_dbf_bytes(data: bytes) -> list[dict]
  parse_excel_bytes(data: bytes) -> dict
  parse_stopa_bytes(data: bytes) -> dict
"""

import io
import struct
import os


# ── DBF parser (bez pliku tymczasowego dla małych danych) ──────────────────

_DBF_ENCODING = "cp1250"


def _read_dbf_from_bytes(data: bytes, encoding: str = _DBF_ENCODING) -> list:
    """Parse raw DBF bytes → list of dicts (same logic as mrpips01_parser._read_dbf_file)."""
    f = io.BytesIO(data)

    # Header
    f.read(1)  # version
    f.read(3)  # date
    num_records = struct.unpack("<I", f.read(4))[0]
    header_size = struct.unpack("<H", f.read(2))[0]
    record_size = struct.unpack("<H", f.read(2))[0]
    f.seek(32)  # skip to field descriptors

    fields = []
    while True:
        raw = f.read(32)
        if not raw or raw[0] == 0x0D:
            break
        name = raw[0:11].decode(encoding, errors="replace").strip("\x00")
        ftype = chr(raw[11])
        length = raw[16]
        fields.append((name, ftype, length))

    f.seek(header_size)
    records = []
    for _ in range(num_records):
        deletion_flag = f.read(1)
        if not deletion_flag:
            break
        if deletion_flag == b"*":  # deleted record
            f.read(record_size - 1)
            continue
        row = {}
        for name, ftype, length in fields:
            raw = f.read(length)
            val = raw.decode(encoding, errors="replace").strip()
            if ftype == "N":
                try:
                    val = int(val) if "." not in val else float(val)
                except ValueError:
                    val = None
            row[name] = val
        records.append(row)
    return records


def parse_dbf_bytes(data: bytes) -> list:
    """Public API: parse DBF file bytes, return list of record dicts."""
    return _read_dbf_from_bytes(data)


# ── Excel parser (wynagrodzenia_claude.xlsx) ───────────────────────────────

def parse_excel_bytes(data: bytes) -> dict:
    """
    Parse wynagrodzenia_claude.xlsx bytes.
    Returns dict with keys: powiaty, struktura_umow, podsumowanie.
    Mirrors logic from extract_data.py load_wynagrodzenia_xlsx().
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)

    def _rows(sheet_name, min_row=1):
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        return list(ws.iter_rows(min_row=min_row, values_only=True))

    # Ark 1: WYN. WG POWIATÓW — wiersz 5+: (ranking, powiat, wyn_ogol, wyn_uop)
    powiaty = []
    for row in _rows("1. WYN. WG POWIATÓW", min_row=5):
        if row[0] is None or not isinstance(row[0], (int, float)):
            continue
        try:
            powiaty.append({
                "ranking": int(row[0]),
                "powiat": str(row[1]).strip() if row[1] is not None else None,
                "wyn_brutto": float(row[2]) if row[2] is not None else None,
                "wyn_uop": float(row[3]) if row[3] is not None else None,
            })
        except (TypeError, ValueError):
            continue

    # Ark 3: STRUKTURA UMÓW — R5 to R73 (rankingi)
    struktura_raw = _rows("3. STRUKTURA UMÓW", min_row=1)

    # Ark 4: PODSUMOWANIE — avg_wyn_uop, top5/bot5 powiaty, top10/bot10 zawody
    podsumowanie_raw = _rows("4. PODSUMOWANIE", min_row=1)

    wb.close()

    return {
        "powiaty": powiaty,
        "struktura_rows": len(struktura_raw),
        "podsumowanie_rows": len(podsumowanie_raw),
        # Raw rows returned for further processing by caller if needed:
        "_struktura": [[c for c in r] for r in struktura_raw],
        "_podsumowanie": [[c for c in r] for r in podsumowanie_raw],
    }


# ── Stopa bezrobocia (GUS BDL) ─────────────────────────────────────────────

def parse_stopa_bytes(data: bytes) -> dict:
    """
    Parse stopa_bezrobocia XLSX bytes (arkusz Tabl.1a).
    Kolumny: WOJ(0), POW(1), Nazwa(2), Bezrob_tys(3), Stopa%(4)

    Zwraca:
      powiaty_maz  — {wgm_str: stopa} dla powiatów mazowieckich (WOJ=14, POW≠00)
      stopa_maz    — stopa woj. mazowieckiego (WOJ=14, POW=00)
      stopa_pl     — stopa dla Polski (WOJ=0, POW=00)
      woj_stopy    — {woj_code: stopa} dla pozostałych 15 województw
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    if "Tabl.1a" not in wb.sheetnames:
        wb.close()
        raise ValueError("Brak arkusza 'Tabl.1a' w pliku XLSX")

    ws = wb["Tabl.1a"]
    powiaty_maz = {}
    stopa_maz = None
    stopa_pl = None
    woj_stopy = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        try:
            woj = str(int(float(str(row[0]).strip())))
        except (ValueError, TypeError):
            continue
        try:
            pow_code = str(int(float(str(row[1]).strip()))).zfill(2)
        except (ValueError, TypeError):
            continue
        try:
            stopa = float(row[4]) if row[4] is not None else None
        except (ValueError, TypeError):
            stopa = None

        if pow_code == "00":
            if woj == "0":
                stopa_pl = stopa
            elif woj == "14":
                stopa_maz = stopa
            elif stopa is not None:
                woj_stopy[woj] = stopa
        elif woj == "14" and stopa is not None:
            powiaty_maz["14" + pow_code] = stopa

    wb.close()
    return {
        "powiaty_maz": powiaty_maz,
        "stopa_maz": stopa_maz,
        "stopa_pl": stopa_pl,
        "woj_stopy": woj_stopy,
    }
