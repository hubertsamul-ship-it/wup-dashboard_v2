"""
WUP Dashboard — skrypt ekstrakcji danych
=========================================
Źródła danych (wszystkie w C:\\Users\\HSamul\\Projekty_AI\\dane\\):

  dane/dbf/{rok}/MP1{RR}{MM}14.dbf
      — dane MRPiPS-01: bezrobotni, zarej., wyrej., kategorie
      — format stary (do 2025): R1-R15; format nowy (2026+): R1-R22
      — struktura identyczna, parser obsługuje oba automatycznie

  dane/stopa_bezrobocia/{rok}/{rok}_{MM}.xlsx
      — stopy bezrobocia wg GUS BDL
      — arkusz Tabl.1a: WOJ(0), POW(1), Nazwa(2), Bezrob_tys(3), Stopa%(4)

  dane/wynagrodzenia/wynagrodzenia_claude.xlsx
      — dane ZUS o wynagrodzeniach (arkusze 1-4)

Użycie:
    python extract_data.py

Wymagane biblioteki:
    pip install openpyxl

Dane wyjściowe (public/data/):
    mrpips_data.json   — bezrobotni wg powiatów, wszystkie dostępne okresy
    wynagr_data.json   — wynagrodzenia wg powiatów, zawodów, struktura umów
"""

import json
import os
import sys

# Dodaj ścieżkę do parsera MRPiPS (ninja/ przy katalogu nadrzędnym)
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'ninja'))

OUT_DIR = os.path.join(os.path.dirname(__file__), 'public', 'data')
os.makedirs(OUT_DIR, exist_ok=True)

# ─── Ścieżki do danych ─────────────────────────────────────────────────────
BASE      = os.path.join(os.path.dirname(__file__), '..')   # Projekty_AI/
DBF_DIR   = os.path.join(BASE, 'dane', 'dbf')
STOPA_DIR = os.path.join(BASE, 'dane', 'stopa_bezrobocia')
WYNAGR_FILE = os.path.join(BASE, 'dane', 'wynagrodzenia', 'wynagrodzenia_claude.xlsx')
ZWOL_DIR    = os.path.join(BASE, 'dane', 'zwolnienia')

# ─── Kody WGM → nazwy powiatów mazowieckich (woj. 14) ─────────────────────
# Mapowanie: ostatnie 2 cyfry WGM = numer arkusza w pliku XLSX MRPiPS
POW_NAMES = {
    '1401': 'Białobrzeski',
    '1402': 'Ciechanowski',
    '1403': 'Garwoliński',
    '1404': 'Gostyniński',
    '1405': 'Grodziski',
    '1406': 'Grójecki',
    '1407': 'Kozienicki',
    '1408': 'Legionowski',
    '1409': 'Lipski',
    '1410': 'Łosicki',
    '1411': 'Makowski',
    '1412': 'Miński',
    '1413': 'Mławski',
    '1414': 'Nowodworski',
    '1415': 'Ostrołęcki',
    '1416': 'Ostrowski',
    '1417': 'Otwocki',
    '1418': 'Piaseczyński',
    '1419': 'Płocki',
    '1420': 'Płoński',
    '1421': 'Pruszkowski',
    '1422': 'Przasnyski',
    '1423': 'Przysuski',
    '1424': 'Pułtuski',
    '1425': 'Radomski',
    '1426': 'Siedlecki',
    '1427': 'Sierpecki',
    '1428': 'Sochaczewski',
    '1429': 'Sokołowski',
    '1430': 'Szydłowiecki',
    # 1431 — brak w rejestrze MRPiPS (nieistniejący powiat)
    '1432': 'Warszawski zachodni',
    '1433': 'Węgrowski',
    '1434': 'Wołomiński',
    '1435': 'Wyszkowski',
    '1436': 'Zwoleński',
    '1437': 'Żuromiński',
    '1438': 'Żyrardowski',
    '1461': 'm. Ostrołęka',
    '1462': 'm. Płock',
    '1463': 'm. Radom',
    '1464': 'm. Siedlce',
    '1465': 'm. Warszawa',
}


# ─── Stopa bezrobocia z GUS ────────────────────────────────────────────────

def load_stopa_xlsx(rok: str, miesiac: str) -> tuple:
    """
    Ładuje stopy bezrobocia rejestrowanego z GUS BDL.
    Plik: dane/stopa_bezrobocia/{rok}/{rok}_{miesiac}.xlsx
    Arkusz Tabl.1a — kolumny: WOJ(0), POW(1), Nazwa(2), Bezrob_tys(3), Stopa%(4)

    Zwraca: (powiaty_dict, stopa_maz, stopa_pl, woj_stopy)
      - powiaty_dict: {wgm_str: stopa} dla powiatów mazowieckich (WOJ=14, POW≠00)
      - stopa_maz:    stopa woj. mazowieckiego (WOJ=14, POW=00)
      - stopa_pl:     stopa dla Polski (WOJ=0, POW=00)
      - woj_stopy:    {woj_code: stopa} dla pozostałych 15 województw (POW=00)
    """
    fpath = os.path.join(STOPA_DIR, rok, f'{rok}_{miesiac}.xlsx')
    if not os.path.exists(fpath):
        print(f"  WARN stopa: brak pliku {os.path.relpath(fpath, BASE)}")
        return {}, None, None, {}, None

    try:
        import openpyxl
    except ImportError:
        print("  WARN: brak openpyxl — pip install openpyxl")
        return {}, None, None, {}, None

    try:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        if 'Tabl.1a' not in wb.sheetnames:
            print(f"  WARN stopa: brak arkusza Tabl.1a w {os.path.basename(fpath)}")
            wb.close()
            return {}, None, None, {}, None

        ws = wb['Tabl.1a']
        result    = {}    # {wgm: stopa} dla powiatów mazowieckich
        stopa_maz = None  # agregat Mazowsza (WOJ=14, POW=00)
        stopa_pl    = None  # agregat Polski   (WOJ=0,  POW=00)
        bezr_pl_tys = None  # bezrobotni PL w tys. (kolumna Bezrob_tys, WOJ=0, POW=00)
        woj_stopy = {}    # {woj_code: stopa} dla pozostałych 15 woj.
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            try:
                woj = str(int(float(str(row[0]).strip())))
            except (ValueError, TypeError):
                continue
            try:
                pow_code = str(int(float(str(row[1]).strip()))).zfill(2)
            except (ValueError, TypeError):
                continue
            try:
                stopa = float(row[4]) if row[4] is not None else None
            except (ValueError, TypeError):
                stopa = None

            if pow_code == '00':
                if woj == '0':
                    stopa_pl = stopa           # wiersz POLSKA
                    try:
                        bezr_pl_tys = float(row[3]) if row[3] is not None else None
                    except (ValueError, TypeError):
                        bezr_pl_tys = None
                elif woj == '14':
                    stopa_maz = stopa          # agregat Mazowsza
                elif stopa is not None:
                    woj_stopy[woj] = stopa     # agregaty pozostałych woj.
            elif woj == '14':
                wgm = '14' + pow_code          # powiaty mazowieckie
                if stopa is not None:
                    result[wgm] = stopa

        wb.close()
        return result, stopa_maz, stopa_pl, woj_stopy, bezr_pl_tys

    except Exception as e:
        print(f"  WARN stopa: błąd czytania {os.path.basename(fpath)}: {e}")
        return {}, None, None, {}, None


# ─── Odkrywanie plików DBF ─────────────────────────────────────────────────

def discover_dbf_files() -> dict:
    """
    Skanuje dane/dbf/{rok}/ w poszukiwaniu plików MRPiPS-01 dla Mazowsza.
    Format nazwy: MP1{RR}{MM}14.dbf  (np. MP1251214.dbf → 2025-12, woj.14)
    Zwraca: dict {'RRRR-MM': ścieżka_pliku} posortowany chronologicznie.
    """
    periods = {}
    if not os.path.exists(DBF_DIR):
        print(f"BRAK katalogu DBF: {DBF_DIR}")
        return periods

    for year_dir in sorted(os.listdir(DBF_DIR)):
        year_path = os.path.join(DBF_DIR, year_dir)
        if not os.path.isdir(year_path):
            continue

        for fname in sorted(os.listdir(year_path)):
            if not fname.upper().endswith('.DBF'):
                continue

            uname = os.path.splitext(fname)[0].upper()
            # Format: MP1 + RR(2) + MM(2) + WW(2)  → MP1251214
            if uname.startswith('MP1') and len(uname) >= 9:
                rok_short = uname[3:5]   # np. '25'
                miesiac   = uname[5:7]   # np. '12'
                woj       = uname[7:9]   # np. '14'
                if woj == '14':
                    period_key = f'20{rok_short}-{miesiac}'
                    periods[period_key] = os.path.join(year_path, fname)

    # Posortuj chronologicznie
    return dict(sorted(periods.items()))


# ─── Ekstrakcja danych MRPiPS ──────────────────────────────────────────────

def extract_mrpips(filename: str, stopa_dict: dict) -> dict:
    """
    Ekstrahuje kluczowe dane z pliku DBF/XLSX MRPiPS-01.
    Obsługuje oba formaty automatycznie:
      - stary (do 2025): D1 T1 NRW=001-020, D1 T2 NRW=021-066 (wyrej=NRW=030)
      - nowy  (2026+):   D1 T1 NRW=001-025, D1 T2 NRW=026-072 (wyrej=NRW=034)

    Pola:
      bezr_razem    — D1 T1 NRW=001 R5 (stan na koniec miesiąca, ogółem)
      bezr_kobiety  — D1 T1 NRW=001 R6
      zarej_razem   — D1 T1 NRW=001 R1 (zarejestrowali się w miesiącu)
      wyrej_razem   — D1 T2 NRW=034/030 R1 (wyłączeni z ewidencji)
      podjeli_prace — D1 T2 NRW=035/031 R1
      oferty_pracy  — D3 T1 NRW=001 R1 (wolne miejsca pracy)
      dlugotrwale   — D5 T1 NRW=006+007 R1 (czas > 12 mies.) — oba formaty
      niepelnosprawni — D1 T1 NRW=023 R5 (tylko nowy format)
      do_30         — D1 T1 NRW=018 R5
      po_50         — D1 T1 NRW=020 R5
      bez_kwalifikacji — D1 T1 NRW=022 R5 (tylko nowy format)
      na_wsi        — D1 T1 NRW=005 R5
      zwolnienia_grupowe — D4 T1 NRW=002 R2
      d5_kobiety    — D5 T1 NRW=001 R2
      d5_czas       — D5 T1 NRW=002-007 R1 [do1, 1-3, 3-6, 6-12, 12-24, pow24]
      d5_wiek       — D5 T1 NRW=008-013 R1 [18-24, 25-34, 35-44, 45-54, 55-59, 60+]
      d5_wyk        — D5 T1 NRW=014-018 R1 [wyższe, pol/śr.zaw, śr.og, zasad, podst]
    """
    try:
        from mrpips01_parser import MRPiPS01
        m = MRPiPS01(filename)
    except ImportError:
        print("  WARN: mrpips01_parser nie znaleziony — sprawdź ścieżkę ninja/")
        return {}
    except Exception as e:
        print(f"  WARN: błąd wczytywania {os.path.basename(filename)}: {e}")
        return {}

    # Wykryj stary format: D1 T1 sięga tylko do NRW=020 (brak NRW=021-025)
    # W starym formacie D1 T2 zaczyna się od NRW=021 (nowy: od NRW=026)
    # Wyrej: stary=NRW=030 R1, nowy=NRW=034 R1
    old_fmt = not any(
        r['DZIAL'] == '1' and r['TABELA'] == '1' and r['NRW'] == '021'
        for r in m.records
    )
    wyrej_nrw   = '030' if old_fmt else '034'
    podjeli_nrw = '031' if old_fmt else '035'

    def wr(new_nrw: str) -> str:
        """NRW przyczyny wyrejestrowania: nowy format vs stary (stały offset -4)."""
        return str(int(new_nrw) - 4).zfill(3) if old_fmt else new_nrw

    wyniki = {}
    for p in m.powiaty:
        # Tylko Mazowieckie (WGM zaczyna się od 14)
        if not str(p).startswith('14'):
            continue

        bezr = m.bezrobotni_na_koniec_miesiaca(p)
        reg  = m.rejestracje_w_miesiacu(p)

        # Wyrej — z poprawnym NRW w zależności od formatu
        wyrej_razem   = m.get_value(p, '1', '2', wyrej_nrw,   'R1')
        podjeli_prace = m.get_value(p, '1', '2', podjeli_nrw, 'R1')

        # Dlugotrwale z D5 (czas > 12 mies.): działa w obu formatach
        dlugotrwale = (m.get_value(p, '5', '1', '006', 'R1') +
                       m.get_value(p, '5', '1', '007', 'R1'))

        # Oferty pracy — D3 T1 NRW=001 R1 (działa w obu formatach)
        oferty_pracy = m.get_value(p, '3', '1', '001', 'R1')

        # D5: Struktura bezrobotnych wg cech
        d5_kobiety = m.get_value(p, '5', '1', '001', 'R2')
        d5_czas = [m.get_value(p, '5', '1', str(i).zfill(3), 'R1') for i in range(2, 8)]
        d5_wiek = [m.get_value(p, '5', '1', str(i).zfill(3), 'R1') for i in range(8, 14)]
        d5_wyk  = [m.get_value(p, '5', '1', str(i).zfill(3), 'R1') for i in range(14, 19)]
        d5_staz = [m.get_value(p, '5', '1', str(i).zfill(3), 'R1') for i in range(19, 26)]

        # Dzieci (D1 T1 NRW=010 R5 — "Mający co najmniej 1 dziecko do 18 lat", oba formaty)
        dzieci           = m.get_value(p, '1', '1', '010', 'R5')

        # Przyczyny wyrejestrowania (D1 T2) — offset -4 w starym formacie
        wyrej_staz        = m.get_value(p, '1', '2', wr('056'), 'R1')
        wyrej_szkolenie   = m.get_value(p, '1', '2', wr('054'), 'R1') + wyrej_staz
        wyrej_niepodjecie = m.get_value(p, '1', '2', wr('060'), 'R1')
        wyrej_brak_kont   = m.get_value(p, '1', '2', wr('063'), 'R1')
        wyrej_dobrowolna  = m.get_value(p, '1', '2', wr('064'), 'R1')
        wyrej_wiek_emet   = m.get_value(p, '1', '2', wr('066'), 'R1')
        wyrej_prawa_emet  = m.get_value(p, '1', '2', wr('067'), 'R1')
        wyrej_inne        = m.get_value(p, '1', '2', wr('069'), 'R1')

        wgm_str = str(p)
        wyniki[wgm_str] = {
            'wgm':              wgm_str,
            'nazwa':            POW_NAMES.get(wgm_str, wgm_str),
            'stopa':            stopa_dict.get(wgm_str),          # z GUS xlsx
            'bezr_razem':       bezr['razem'],
            'bezr_kobiety':     bezr['kobiety'],
            'zarej_razem':      reg['razem'],
            'wyrej_razem':      wyrej_razem,
            'podjeli_prace':    podjeli_prace,
            'oferty_pracy':     oferty_pracy,
            'dlugotrwale':      dlugotrwale,
            'niepelnosprawni':  m.get_value(p, '1', '1', '023', 'R5'),
            'do_30':            m.get_value(p, '1', '1', '018', 'R5'),
            'po_50':            m.get_value(p, '1', '1', '020', 'R5'),
            'bez_kwalifikacji': m.get_value(p, '1', '1', '022', 'R5'),
            'na_wsi':           m.get_value(p, '1', '1', '005', 'R5'),
            'zwolnienia_grupowe': m.get_value(p, '4', '1', '002', 'R2'),
            'd5_kobiety':       d5_kobiety,
            'd5_czas':          d5_czas,
            'd5_wiek':          d5_wiek,
            'd5_wyk':           d5_wyk,
            'dzieci':           dzieci,
            'wyrej_szkolenie':   wyrej_szkolenie,
            'wyrej_niepodjecie': wyrej_niepodjecie,
            'wyrej_brak_kont':   wyrej_brak_kont,
            'wyrej_dobrowolna':  wyrej_dobrowolna,
            'wyrej_wiek_emet':   wyrej_wiek_emet,
            'wyrej_prawa_emet':  wyrej_prawa_emet,
            'wyrej_inne':        wyrej_inne,
            'd5_staz':          d5_staz,
            'rok':              m.rok,
            'miesiac':          m.miesiac,
        }

    return wyniki


# ─── Ekstrakcja wynagrodzeń ────────────────────────────────────────────────

def extract_wynagrodzenia(filename: str) -> dict:
    """Ekstrahuje dane z arkuszy wynagrodzeniowych (arkusze 1-4)."""
    try:
        import openpyxl
    except ImportError:
        print("  WARN: openpyxl nie zainstalowane — pip install openpyxl")
        return {}

    try:
        wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
    except Exception as e:
        print(f"  WARN: Nie można wczytać {os.path.basename(filename)}: {e}")
        return {}

    result = {
        'powiaty': [],
        'zawody': [],
        'struktura_umow': [],
        'podsumowanie': {},
    }

    # Ark 1: WYN. WG POWIATÓW — cols: 0=rank, 1=powiat, 2=wyn_ogółem, 3=wyn≤2lat, 4=wyn>2lat, 5=pracujący
    sname1 = '1. WYN. WG POWIATÓW'
    if sname1 in wb.sheetnames:
        ws = wb[sname1]
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not row[0] or not row[1]:
                continue
            try:
                result['powiaty'].append({
                    'powiat':        str(row[1]).strip(),
                    'wynagrodzenie': float(row[2]) if row[2] is not None else 0,
                    'wyn_krotki':    float(row[3]) if len(row) > 3 and row[3] is not None else 0,
                    'wyn_dlugi':     float(row[4]) if len(row) > 4 and row[4] is not None else 0,
                    'pracujacy':     int(float(row[5])) if len(row) > 5 and row[5] is not None else 0,
                })
            except (TypeError, ValueError):
                pass
    else:
        print(f"  WARN: brak arkusza '{sname1}' — dostępne: {wb.sheetnames}")

    # Ark 2: WYN. WG ZAWODÓW — cols: 0=rank, 1=zawód, 2=wyn_ogółem, 3=wyn≤2lat, 4=wyn>2lat, 5=premia%
    sname2 = '2. WYN. WG ZAWODÓW'
    if sname2 in wb.sheetnames:
        ws = wb[sname2]
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not row[0] or not row[1]:
                continue
            try:
                result['zawody'].append({
                    'zawod':         str(row[1]).strip(),
                    'wynagrodzenie': float(row[2]) if row[2] is not None else 0,
                    'premia_pct':    float(row[5]) if len(row) > 5 and row[5] is not None else 0,
                })
            except (TypeError, ValueError):
                pass

    # Ark 3: STRUKTURA UMÓW — cols: 0=rank, 1=powiat, 2=pracujący, 3=nowozatr, 4=pct_uop, 5=pct_cywil
    sname3 = '3. STRUKTURA UMÓW'
    if sname3 in wb.sheetnames:
        ws = wb[sname3]
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not row[1]:
                continue
            try:
                result['struktura_umow'].append({
                    'powiat':    str(row[1]).strip(),
                    'pracujacy': int(float(row[2])) if row[2] is not None else 0,
                    'pct_uop':   float(row[4]) if len(row) > 4 and row[4] is not None else 0,
                    'pct_cywil': float(row[5]) if len(row) > 5 and row[5] is not None else 0,
                })
            except (TypeError, ValueError):
                pass

    # Ark 3 c.d.: pracujacy — KPI (R5) + rankingi (R52-73)
    if sname3 in wb.sheetnames:
        ws3 = wb[sname3]
        def _row(r):
            return next(ws3.iter_rows(min_row=r, max_row=r, values_only=True), [None]*12)
        def _pairs(r1, r2):
            left, right = [], []
            for row in ws3.iter_rows(min_row=r1, max_row=r2, values_only=True):
                if row[1] is not None: left.append({'label': str(row[1]), 'value': row[2]})
                if row[5] is not None: right.append({'label': str(row[5]), 'value': row[6]})
            return left, right
        k = _row(5)
        p_top, p_bot  = _pairs(52, 56)
        z_prac, z_uop = _pairs(61, 65)
        _,      z_dg  = _pairs(69, 73)

        # Nowa tabela: TOP 5 ZAWODÓW Z DOMINACJĄ UMÓW CYWILNOPRAWNYCH (R84-90, kol A/B/C)
        c_pow = []
        for row in ws3.iter_rows(min_row=84, max_row=90, values_only=True):
            rank = row[0] if len(row) > 0 else None
            lbl  = row[1] if len(row) > 1 else None
            val  = row[2] if len(row) > 2 else None
            if not isinstance(rank, (int, float)) or lbl is None:
                continue  # pomiń nagłówek i puste wiersze
            # Excel przechowuje % jako ułamek dziesiętny (np. 0.8431) → przelicz na 0-100
            if isinstance(val, float) and 0 < val <= 1.0:
                val = round(val * 100, 2)
            c_pow.append({'label': str(lbl).strip(), 'value': val})

        result['pracujacy'] = {
            'razem': int(k[2] or 0),
            'pct_uop': k[4], 'n_uop':   int(k[8] or 0),
            'pct_cywil': k[5], 'n_cywil': int(k[10] or 0),
            'pct_dg':  k[6],  'n_dg':    int(k[11] or 0),
            'top5_pracujacy': p_top,  'bot5_pracujacy': p_bot,
            'top5_zawody_pracujacy': z_prac, 'top5_zawody_uop': z_uop,
            'top5_pow_cywil': c_pow,  'top5_zawody_dg': z_dg,
        }

    # Ark 4: PODSUMOWANIE — KPI (R5) + rankingi powiatów (R17-29) + zawodów (R33-55)
    sname4 = '4. PODSUMOWANIE'
    if sname4 in wb.sheetnames:
        ws4 = wb[sname4]
        # Wczytaj wszystkie wiersze jednorazowo (read_only — single pass)
        rows4 = {i: row for i, row in enumerate(ws4.iter_rows(values_only=True), 1)}

        def g4(r, c):
            row = rows4.get(r, ())
            return row[c] if c < len(row) else None

        def list4(r1, r2, with_szac=False):
            items = []
            for r in range(r1, r2 + 1):
                lbl = g4(r, 1)
                val = g4(r, 2)
                if lbl is not None and val is not None:
                    try:
                        item = {'label': str(lbl).strip(), 'value': round(float(val), 2)}
                        if with_szac:
                            szac = g4(r, 3)
                            if szac is not None:
                                item['szac_uop'] = int(float(szac))
                        items.append(item)
                    except (TypeError, ValueError):
                        pass
            return items

        avg_wyn_uop = None
        try:
            v5 = g4(5, 1)
            if v5 is not None:
                avg_wyn_uop = round(float(v5), 2)
        except (TypeError, ValueError):
            pass

        result['podsumowanie'] = {'avg_wyn_uop': avg_wyn_uop}
        result['wynagrodzenia_kpi'] = {
            'avg_wyn_uop':        avg_wyn_uop,
            'top5_pow_wyn':       list4(17, 21),
            'bot5_pow_wyn':       list4(25, 29),
            'top10_zawody_best':  list4(33, 42, with_szac=True),
            'top10_zawody_worst': list4(46, 55, with_szac=True),
        }

    wb.close()
    return result


# ─── Zwolnienia grupowe z xlsx ──────────────────────────────────────────────

def _roman_to_int(s: str):
    vals = {'I':1,'II':2,'III':3,'IV':4,'V':5,'VI':6,'VII':7,'VIII':8,'IX':9,'X':10,'XI':11,'XII':12}
    return vals.get(s.upper())


def extract_zwolnienia() -> dict:
    """
    Wczytuje dane o zwolnieniach grupowych z plików xlsx.
    Ścieżka: dane/zwolnienia/{YYYY}/{RZ}_{YYYY}.xlsx
    Arkusz 'dane', wiersze od 8:
      kol B (idx 1)  = Powiat (pusty → wiersz sumaryczny, pomijamy)
      kol F (idx 5)  = PKD
      kol G (idx 6)  = zgłoszenia
      kol I (idx 8)  = wypowiedzenia zmieniające (kumulatywne per firma)
      kol K (idx 10) = faktyczne zwolnienia (kumulatywne per firma)
      kol L (idx 11) = monitorowane (kumulatywne per firma)
      kol O (idx 14) = Data zgłoszenia informacji do PUP

    Logika deduplicji:
      - pomijamy wiersze bez powiatu (sumaryczne)
      - liczymy wiersz tylko w miesiącu pasującym do daty zgłoszenia
      - wiersze bez daty: liczymy w bieżącym pliku, tylko dla nowych REGON-ów
      - zgłoszenia: tylko przy pierwszym wystąpieniu REGON-u
      - wyd/fakt/mon: kumulatywne → liczymy deltę (przyrost) między miesiącami
    """
    import re as _re
    from datetime import datetime as _dt

    MIESIACE = ['Sty','Lut','Mar','Kwi','Maj','Cze','Lip','Sie','Wrz','Paź','Lis','Gru']

    def parse_date(val):
        if isinstance(val, _dt): return (val.year, val.month)
        if isinstance(val, str):
            m = _re.search(r'(\d{1,2})\.(\d{1,2})\.(\d{4})', val.strip())
            if m: return (int(m.group(3)), int(m.group(2)))
        return None

    try:
        import openpyxl
    except ImportError:
        print("  WARN: brak openpyxl — pip install openpyxl")
        return {}

    if not os.path.isdir(ZWOL_DIR):
        print(f"  WARN zwolnienia: brak katalogu {os.path.relpath(ZWOL_DIR, BASE)}")
        return {}

    # Zbierz i posortuj wszystkie pliki chronologicznie
    all_files = []
    for rok_dir in os.listdir(ZWOL_DIR):
        rok_path = os.path.join(ZWOL_DIR, rok_dir)
        if not os.path.isdir(rok_path): continue
        try: rok = int(rok_dir)
        except ValueError: continue
        for fname in os.listdir(rok_path):
            if not fname.lower().endswith('.xlsx'): continue
            mc_int = _roman_to_int(fname.upper().replace('.XLSX','').split('_')[0])
            if mc_int is None: continue
            all_files.append((rok, mc_int, os.path.join(rok_path, fname)))
    all_files.sort()

    months_data = {}  # (rok, mc) → {zgl, wyd, fakt, mon, pkd, pkd_fakt}

    for (rok, mc_int, fpath) in all_files:
        key = (rok, mc_int)
        months_data[key] = {'zgl': 0, 'wyd': 0, 'fakt': 0, 'mon': 0, 'pkd': {}, 'pkd_fakt': {}}

        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            ws = wb['dane'] if 'dane' in wb.sheetnames else wb.active

            # Odczytaj właściwy miesiąc (D2) i rok (E2) z nagłówka arkusza
            for hdr in ws.iter_rows(min_row=2, max_row=2, min_col=4, max_col=5, values_only=True):
                d2, e2 = hdr[0], hdr[1]
                if isinstance(d2, (int, float)) and isinstance(e2, (int, float)) \
                        and 1 <= int(d2) <= 12 and int(e2) >= 2020:
                    new_key = (int(e2), int(d2))
                    if new_key != key:
                        del months_data[key]
                        key = new_key
                    months_data[key] = {'zgl': 0, 'wyd': 0, 'fakt': 0, 'mon': 0, 'pkd': {}, 'pkd_fakt': {}}
                break

            for row in ws.iter_rows(min_row=8, values_only=True):
                if not row or len(row) < 10: continue

                # Pomiń wiersze sumaryczne (pusty powiat = kol B = idx 1)
                powiat = row[1]
                if powiat is None or (isinstance(powiat, str) and not powiat.strip()):
                    continue

                v_pkd  = row[5]  if len(row) > 5  else None
                v_zgl  = int(row[6])  if len(row) > 6  and isinstance(row[6],  (int, float)) and row[6]  > 0 else 0
                v_wyd  = int(row[8])  if len(row) > 8  and isinstance(row[8],  (int, float)) and row[8]  > 0 else 0
                v_fakt = int(row[10]) if len(row) > 10 and isinstance(row[10], (int, float)) and row[10] > 0 else 0
                v_mon  = int(row[11]) if len(row) > 11 and isinstance(row[11], (int, float)) and row[11] > 0 else 0

                # Zlicz ten wiersz (miesiąc określa nagłówek D2/E2, data z kol. O nie jest używana)
                months_data[key]['zgl']  += v_zgl
                months_data[key]['wyd']  += v_wyd
                months_data[key]['fakt'] += v_fakt
                months_data[key]['mon']  += v_mon

                # PKD
                pkd2 = (v_pkd[:2].strip() if isinstance(v_pkd, str) and len(v_pkd) >= 2 else None)
                if pkd2:
                    if v_zgl > 0:  months_data[key]['pkd'][pkd2]      = months_data[key]['pkd'].get(pkd2, 0)      + v_zgl
                    if v_fakt > 0: months_data[key]['pkd_fakt'][pkd2] = months_data[key]['pkd_fakt'].get(pkd2, 0) + v_fakt

            wb.close()
            d = months_data[key]
            print(f"  zwolnienia {key[0]}-{key[1]:02d}: zgl={d['zgl']:,}  wyd={d['wyd']:,}  fakt={d['fakt']:,}  mon={d['mon']:,}")

        except Exception as e:
            print(f"  WARN zwolnienia: {os.path.basename(fpath)}: {e}")

    if not months_data:
        return {}

    sorted_keys = sorted(months_data.keys())
    last_13 = sorted_keys[-14:]

    trend_13m = []
    for (rok, mc) in last_13:
        d = months_data[(rok, mc)]
        trend_13m.append({
            'label': f"{MIESIACE[mc-1]} {str(rok)[2:]}",
            'zgl': d['zgl'], 'wyd': d['wyd'], 'fakt': d['fakt'], 'mon': d['mon'],
        })

    cur = months_data[sorted_keys[-1]]
    prv = months_data[sorted_keys[-2]] if len(sorted_keys) >= 2 else cur

    # PKD: agregat z ostatnich 12 miesięcy (zgłoszenia + faktyczne)
    pkd_agg = {}
    pkd_agg_fakt = {}
    for (rok, mc) in sorted_keys[-12:]:
        for kod, cnt in months_data[(rok, mc)]['pkd'].items():
            pkd_agg[kod] = pkd_agg.get(kod, 0) + cnt
        for kod, cnt in months_data[(rok, mc)].get('pkd_fakt', {}).items():
            pkd_agg_fakt[kod] = pkd_agg_fakt.get(kod, 0) + cnt
    pkd_top10      = [{'pkd': k, 'n': v} for k, v in sorted(pkd_agg.items(),      key=lambda x: -x[1])[:10]]
    pkd_top10_fakt = [{'pkd': k, 'n': v} for k, v in sorted(pkd_agg_fakt.items(), key=lambda x: -x[1])[:10]]

    return {
        'zgl_cur':   cur['zgl'],
        'zgl_delta': cur['zgl'] - prv['zgl'],
        'fakt_cur':  cur['fakt'],
        'fakt_delta': cur['fakt'] - prv['fakt'],
        'mon_cur':   cur['mon'],
        'mon_delta': cur['mon'] - prv['mon'],
        'wyd_cur':   cur['wyd'],
        'wyd_delta': cur['wyd'] - prv['wyd'],
        'trend_13m': trend_13m,
        'pkd_top10': pkd_top10,
        'pkd_top10_fakt': pkd_top10_fakt,
    }


# ─── Budowanie dashboard_final.json ────────────────────────────────────────

def build_dashboard_final(mrpips_data: dict, wynagr_data: dict, zwolnienia_data: dict = None) -> dict:
    """
    Oblicza wszystkie wartości prezentowane w dashboardzie i zapisuje je
    jako dashboard_final.json. React tylko wyświetla — nie liczy.
    """
    import datetime

    # Posortowane okresy (chronologicznie): ['2023-01', ..., '2026-01']
    periods = sorted(p for p in mrpips_data if not p.startswith('_'))
    if len(periods) < 2:
        print("  WARN build_dashboard_final: zbyt mało okresów")
        return {}

    cur = periods[-1]   # '2026-02' (najnowszy DBF)
    prv = periods[-2]   # '2026-01'

    MIESIACE = ['Sty','Lut','Mar','Kwi','Maj','Cze','Lip','Sie','Wrz','Paź','Lis','Gru']

    def trend_label(period):
        rok, mc = period.split('-')
        return f"{MIESIACE[int(mc)-1]} {rok[2:]}"

    def powiaty(p):
        """Lista słowników powiatów dla okresu p (bez _meta)."""
        return [v for k, v in mrpips_data[p].items()
                if k != '_meta' and isinstance(v, dict) and v.get('wgm')]

    def s(p, field):
        """Suma pola field dla wszystkich powiatów w okresie p."""
        return sum((v.get(field) or 0) for v in powiaty(p))

    def meta(p, field):
        return mrpips_data[p].get('_meta', {}).get(field)

    # ── Fallback stopy: GUS publikuje z opóźnieniem — szukaj ostatniego okresu ze stopą ──
    stopa_cur_period = next(
        (p for p in reversed(periods) if meta(p, 'stopa_maz') is not None),
        prv
    )
    stopa_prev_period = next(
        (p for p in reversed(periods) if p != stopa_cur_period and meta(p, 'stopa_maz') is not None),
        None
    )
    if stopa_cur_period != cur:
        print(f"  INFO stopa: brak stopy dla {cur}, używam {stopa_cur_period}")

    # ── Agregaty bieżący vs poprzedni ───────────────────────────────────────
    bezr_cur   = s(cur, 'bezr_razem')
    bezr_prev  = s(prv, 'bezr_razem')
    wyrej_cur  = s(cur, 'wyrej_razem')
    wyrej_prev = s(prv, 'wyrej_razem')
    zarej_cur  = s(cur, 'zarej_razem')
    zarej_prev = s(prv, 'zarej_razem')
    oferty_cur = s(cur, 'oferty_pracy')
    oferty_prev= s(prv, 'oferty_pracy')
    stopa_cur  = meta(stopa_cur_period,  'stopa_maz')
    stopa_prev = meta(stopa_prev_period, 'stopa_maz') if stopa_prev_period else None

    stopa_delta = (round(stopa_cur - stopa_prev, 1)
                   if stopa_cur is not None and stopa_prev is not None else None)

    stopa_pl_cur  = meta(stopa_cur_period,  'stopa_pl')
    stopa_pl_prev = meta(stopa_prev_period, 'stopa_pl') if stopa_prev_period else None
    stopa_pl_delta = (round(stopa_pl_cur - stopa_pl_prev, 1)
                      if stopa_pl_cur is not None and stopa_pl_prev is not None else None)

    bezr_pl_tys_cur  = meta(stopa_cur_period,  'bezr_pl_tys')
    bezr_pl_tys_prev = meta(stopa_prev_period, 'bezr_pl_tys') if stopa_prev_period else None
    bezr_pl_cur  = round(bezr_pl_tys_cur  * 1000) if bezr_pl_tys_cur  is not None else None
    bezr_pl_prev = round(bezr_pl_tys_prev * 1000) if bezr_pl_tys_prev is not None else None
    bezr_pl_delta = (bezr_pl_cur - bezr_pl_prev) if (bezr_pl_cur is not None and bezr_pl_prev is not None) else None

    # ── Lookup stopy powiatów z okresu ze stopą (fallback dla powiaty/mapy) ──
    stopa_pow_lookup = {
        wgm: v.get('stopa')
        for wgm, v in mrpips_data[stopa_cur_period].items()
        if wgm != '_meta' and isinstance(v, dict) and v.get('stopa') is not None
    }

    def get_stopa_pow(wgm):
        """Stopa dla powiatu — z cur jeśli dostępna, fallback z stopa_cur_period."""
        v = mrpips_data[cur].get(wgm, {})
        return v.get('stopa') if v.get('stopa') is not None else stopa_pow_lookup.get(wgm)

    # ── Kategorie bezrobotnych ───────────────────────────────────────────────
    def kategoria(label, field):
        n = s(cur, field)
        pct = round(n / bezr_cur * 100, 1) if bezr_cur else 0
        return {'label': label, 'n': n, 'pct': pct}

    kategorie = sorted([
        kategoria('Zamieszkali na wsi',          'na_wsi'),
        kategoria('Długotrwale bezrobotni',      'dlugotrwale'),
        kategoria('Bez kwalifikacji zawodowych', 'bez_kwalifikacji'),
        kategoria('Powyżej 50 roku życia',       'po_50'),
        kategoria('Z dzieckiem do 18 lat',       'dzieci'),
        kategoria('Do 30 roku życia',            'do_30'),
        kategoria('Niepełnosprawni',             'niepelnosprawni'),
    ], key=lambda x: x['n'], reverse=True)

    # ── D5: agregaty cech bezrobotnych ───────────────────────────────────────
    pow_cur = powiaty(cur)

    def d5_sum(field, length):
        result = []
        for i in range(length):
            total = 0
            for v in pow_cur:
                lst = v.get(field) or []
                if isinstance(lst, list) and i < len(lst):
                    total += lst[i] or 0
            result.append(total)
        return result

    d5_kobiety = sum((v.get('d5_kobiety') or 0) for v in pow_cur)
    d5_czas    = d5_sum('d5_czas', 6)
    d5_wiek    = d5_sum('d5_wiek', 6)
    d5_wyk     = d5_sum('d5_wyk',  5)
    d5_staz    = d5_sum('d5_staz', 7)

    # ── Trendy 37 miesięcy (pulpit + bezrobotni) ────────────────────────────
    trend_37m = [
        {
            'okres': p,
            'label': trend_label(p),
            'bezr':  s(p, 'bezr_razem'),
            'zarej': s(p, 'zarej_razem'),
            'wyrej': s(p, 'wyrej_razem'),
            'stopa': meta(p, 'stopa_maz'),
        }
        for p in periods
    ]

    # ── Rankingi powiatów (stopa) ────────────────────────────────────────────
    pow_with_stopa = [
        {'wgm': v['wgm'], 'nazwa': v['nazwa'], 'stopa': get_stopa_pow(v['wgm'])}
        for v in pow_cur if get_stopa_pow(v['wgm']) is not None
    ]
    pow_sorted = sorted(pow_with_stopa, key=lambda x: x['stopa'], reverse=True)

    def to_bar(items):
        """Konwertuj [{ wgm, nazwa, stopa }] na format HorizontalBar [{ label, value }]."""
        return [{'label': d['nazwa'], 'value': d['stopa']} for d in items]

    # ── Przyczyny wyrejestrowania — bieżący miesiąc ─────────────────────────
    def wr_reason(label, field):
        n = s(cur, field)
        pct = round(n / wyrej_cur * 100, 1) if wyrej_cur else 0
        return {'label': label, 'n': n, 'pct': pct}

    wyrej_reasons = sorted([
        wr_reason('Podjęcie pracy',          'podjeli_prace'),
        wr_reason('Brak kontaktu z PUP',     'wyrej_brak_kont'),
        wr_reason('Niepodjęcie/przerwanie',  'wyrej_niepodjecie'),
        wr_reason('Dobrowolna rezygnacja',   'wyrej_dobrowolna'),
        wr_reason('Szkolenie i staż',        'wyrej_szkolenie'),
        wr_reason('Wiek/prawa emery.',       'wyrej_wiek_emet'),
        wr_reason('Inne przyczyny',          'wyrej_inne'),
    ], key=lambda x: x['n'], reverse=True)

    # ── Trend stopy Mazowsza — ostatnie 13 miesięcy ─────────────────────────
    last_13 = periods[-13:] if len(periods) >= 13 else periods
    trend_maz_13m = [
        {'okres': p, 'label': trend_label(p), 'stopa': meta(p, 'stopa_maz')}
        for p in last_13
    ]

    # ── Trend stopy Polski — ostatnie 13 miesięcy (GUS BDL) ─────────────────
    trend_pl_13m = [
        {'label': trend_label(p), 'stopa': meta(p, 'stopa_pl')}
        for p in last_13
    ]

    # ── Stopa bezrobocia — wszystkie 16 województw (bieżący miesiąc) ────────
    WOJ_NAMES = {
        '2':  'Dolnośląskie',  '4':  'Kujawsko-pom.',
        '6':  'Lubelskie',     '8':  'Lubuskie',
        '10': 'Łódzkie',       '12': 'Małopolskie',
        '14': 'Mazowieckie',   '16': 'Opolskie',
        '18': 'Podkarpackie',  '20': 'Podlaskie',
        '22': 'Pomorskie',     '24': 'Śląskie',
        '26': 'Świętokrzyskie','28': 'Warmińsko-maz.',
        '30': 'Wielkopolskie', '32': 'Zachodniopom.',
    }
    woj_stopy_cur = dict(meta(stopa_cur_period, 'woj_stopy') or {})
    woj_stopy_cur['14'] = stopa_cur   # Mazowieckie z agregatu GUS
    woj_stopa_list = sorted(
        [{'n': WOJ_NAMES[k], 's': v}
         for k, v in woj_stopy_cur.items()
         if k in WOJ_NAMES and v is not None],
        key=lambda x: x['s'], reverse=True
    )

    # ── Mapa Mazowsza ────────────────────────────────────────────────────────
    mapa_maz = [
        {'wgm': v['wgm'], 'nazwa': v['nazwa'],
         'stopa': get_stopa_pow(v['wgm']), 'bezr_razem': v.get('bezr_razem')}
        for v in pow_cur
    ]

    # ── Powiaty lista z trendem 13 mies. + pełne dane na wzór Bezrobotni ────
    def normalize_name(s):
        return s.lower().replace('m. ', '').replace('m.st. ', '').replace('powiat ', '').replace('.', '').strip()

    wynagr_pow_lookup = {}
    for pw in wynagr_data.get('powiaty', []):
        key = normalize_name(pw['powiat'])
        wynagr_pow_lookup[key] = pw

    def find_wyn(nazwa):
        return wynagr_pow_lookup.get(normalize_name(nazwa), {})

    pow_lista = []
    for wgm, nazwa in POW_NAMES.items():
        trend_s = [mrpips_data[p].get(wgm, {}).get('stopa') for p in last_13]
        trend_z = [mrpips_data[p].get(wgm, {}).get('zarej_razem') for p in last_13]
        trend_w = [mrpips_data[p].get(wgm, {}).get('wyrej_razem') for p in last_13]
        v = mrpips_data[cur].get(wgm, {})
        bezr = v.get('bezr_razem') or 0
        wyrej = v.get('wyrej_razem') or 0
        wyn = find_wyn(nazwa)

        def kat(label, field):
            n = v.get(field) or 0
            return {'label': label, 'n': n, 'pct': round(n / bezr * 100, 1) if bezr else 0}

        def wr_r(label, field):
            n = v.get(field) or 0
            return {'label': label, 'n': n, 'pct': round(n / wyrej * 100, 1) if wyrej else 0}

        pow_lista.append({
            'wgm':   wgm,
            'nazwa': nazwa,
            # KPI
            'stopa':           get_stopa_pow(wgm),
            'bezr_razem':      v.get('bezr_razem'),
            'bezr_kobiety':    v.get('bezr_kobiety'),
            'zarej_razem':     v.get('zarej_razem'),
            'wyrej_razem':     v.get('wyrej_razem'),
            'oferty_pracy':    v.get('oferty_pracy'),
            'podjeli_prace':   v.get('podjeli_prace'),
            # Kategorie
            'kategorie': sorted([
                kat('Zamieszkali na wsi',          'na_wsi'),
                kat('Długotrwale bezrobotni',      'dlugotrwale'),
                kat('Bez kwalifikacji zawodowych', 'bez_kwalifikacji'),
                kat('Powyżej 50 roku życia',       'po_50'),
                kat('Z dzieckiem do 18 lat',       'dzieci'),
                kat('Do 30 roku życia',            'do_30'),
                kat('Niepełnosprawni',             'niepelnosprawni'),
            ], key=lambda x: x['n'], reverse=True),
            # Charakterystyka D5
            'd5_kobiety': v.get('d5_kobiety'),
            'd5_czas':    v.get('d5_czas'),
            'd5_wiek':    v.get('d5_wiek'),
            'd5_wyk':     v.get('d5_wyk'),
            'd5_staz':    v.get('d5_staz'),
            # Przyczyny wyrejestrowania
            'wyrej_reasons': sorted([
                wr_r('Podjęcie pracy',         'podjeli_prace'),
                wr_r('Brak kontaktu z PUP',    'wyrej_brak_kont'),
                wr_r('Niepodjęcie/przerwanie', 'wyrej_niepodjecie'),
                wr_r('Dobrowolna rezygnacja',  'wyrej_dobrowolna'),
                wr_r('Szkolenie i staż',       'wyrej_szkolenie'),
                wr_r('Wiek/prawa emery.',      'wyrej_wiek_emet'),
                wr_r('Inne przyczyny',         'wyrej_inne'),
            ], key=lambda x: x['n'], reverse=True),
            # Trendy 13 mies.
            'trend_stopa_13m': trend_s,
            'trend_zarej_13m': trend_z,
            'trend_wyrej_13m': trend_w,
            # Wynagrodzenia (ZUS ark. 1)
            'wyn_brutto':    wyn.get('wynagrodzenie'),
            'wyn_pracujacy': wyn.get('pracujacy'),
        })

    # ── Wynagrodzenia: maz_avg z pre-computed wynagrodzenia_kpi ─────────────
    maz_avg = (wynagr_data.get('wynagrodzenia_kpi') or {}).get('avg_wyn_uop')

    # ── Meta: ostatni dzień okresu ──────────────────────────────────────────
    rok, mc = cur.split('-')
    import calendar
    last_day_of_month = calendar.monthrange(int(rok), int(mc))[1]
    data_akt = f"{rok}-{mc}-{last_day_of_month:02d}"

    return {
        'meta': {
            'okres':                  cur,
            'poprzedni_okres':        prv,
            'stopa_okres':            stopa_cur_period,
            'stopa_poprzedni_okres':  stopa_prev_period,
            'data_aktualizacji':      data_akt,
        },
        'pulpit': {
            'bezr_razem':      bezr_cur,
            'bezr_delta':      bezr_cur - bezr_prev,
            'stopa_maz':       stopa_cur,
            'stopa_maz_delta': stopa_delta,
            'stopa_pl':        stopa_pl_cur if stopa_pl_cur is not None else 5.4,
            'stopa_pl_delta':  stopa_pl_delta,
            'bezr_pl':         bezr_pl_cur,
            'bezr_pl_delta':   bezr_pl_delta,
            'trend_37m':       trend_37m,
            'mapa_maz':        mapa_maz,
        },
        'bezrobotni': {
            'bezr_razem':  bezr_cur,
            'bezr_delta':  bezr_cur - bezr_prev,
            'wyrej_razem': wyrej_cur,
            'wyrej_delta': wyrej_cur - wyrej_prev,
            'zarej_razem': zarej_cur,
            'zarej_delta': zarej_cur - zarej_prev,
            'oferty_razem': oferty_cur,
            'oferty_delta': oferty_cur - oferty_prev,
            'kategorie':   kategorie,
            'wyrej_reasons': wyrej_reasons,
            'trend_13m': [
                {'label': t['label'], 'zarej': t['zarej'], 'wyrej': t['wyrej']}
                for t in trend_37m[-13:]
            ],
            'charakterystyka': {
                'kobiety':   d5_kobiety,
                'mezczyzni': bezr_cur - d5_kobiety,
                'czas':      d5_czas,
                'wiek':      d5_wiek,
                'wyk':       d5_wyk,
                'staz':      d5_staz,
            },
        },
        'stopa': {
            'stopa_pl':        stopa_pl_cur if stopa_pl_cur is not None else 5.4,
            'stopa_pl_delta':  stopa_pl_delta,
            'stopa_maz':       stopa_cur,
            'stopa_maz_delta': stopa_delta,
            'pow_max':         pow_sorted[0]  if pow_sorted else None,
            'pow_min':         pow_sorted[-1] if pow_sorted else None,
            'pow_top5':        to_bar(pow_sorted[:5]),
            'pow_bot5':        to_bar(pow_sorted[-5:]),
            'trend_maz_13m':   trend_maz_13m,
            'woj_stopa':       woj_stopa_list,
            'trend_pl_13m':    trend_pl_13m,
        },
        'wynagrodzenia': {
            'maz_avg':       maz_avg,
            'powiaty':       wynagr_data.get('powiaty', []),
            'zawody':        wynagr_data.get('zawody', []),
            'struktura_umow': wynagr_data.get('struktura_umow', []),
            'podsumowanie':  wynagr_data.get('podsumowanie', {}),
        },
        'pracujacy':         wynagr_data.get('pracujacy', {}),
        'wynagrodzenia_kpi': wynagr_data.get('wynagrodzenia_kpi', {}),
        'powiaty_lista': pow_lista,
        'zwolnienia': zwolnienia_data or {},
    }


# ─── Główna funkcja ────────────────────────────────────────────────────────

def main():
    print("=" * 65)
    print("WUP Dashboard — ekstrakcja danych")
    print(f"Źródło DBF:   {os.path.relpath(DBF_DIR,   BASE)}")
    print(f"Źródło stopa: {os.path.relpath(STOPA_DIR, BASE)}")
    print("=" * 65)

    # ── 1. Odkryj pliki DBF ──────────────────────────────────────────────
    dbf_files = discover_dbf_files()
    if not dbf_files:
        print("BŁĄD: Nie znaleziono żadnych plików DBF w dane/dbf/")
        return

    print(f"\nZnaleziono {len(dbf_files)} plików DBF:")
    for period, path in dbf_files.items():
        print(f"  {period}  {os.path.basename(path)}")

    # ── 2. Wczytaj dane MRPiPS + stopy GUS ─────────────────────────────
    print(f"\nEkstrakcja MRPiPS-01:")
    mrpips_out = {}

    for period, fpath in dbf_files.items():
        rok, miesiac = period.split('-')
        stopa_dict, stopa_maz, stopa_pl, woj_stopy, bezr_pl_tys = load_stopa_xlsx(rok, miesiac)
        n_stopa = len(stopa_dict)

        print(f"  {period}  {os.path.basename(fpath)} ...", end='  ')
        data = extract_mrpips(fpath, stopa_dict)
        if data:
            # Agregat województwa — stopa GUS (POW=00)
            data['_meta'] = {
                'stopa_maz':  stopa_maz,   # stopa bezrobocia Mazowsza wg GUS (%)
                'stopa_pl':   stopa_pl,    # stopa bezrobocia Polski   wg GUS (%)
                'bezr_pl_tys': bezr_pl_tys, # bezrobotni PL w tys. (np. 934.1)
                'woj_stopy':  woj_stopy,   # {woj_code: stopa} dla 15 pozostałych woj.
            }
            mrpips_out[period] = data
            maz_str = f"{stopa_maz}%" if stopa_maz is not None else "brak"
            print(f"OK ({len(data) - 1} powiatów, stopa woj.: {maz_str}, powiaty: {n_stopa} pkt.)")
        else:
            print("SKIP")

    # ── 3. Zapisz mrpips_data.json ───────────────────────────────────────
    mrpips_path = os.path.join(OUT_DIR, 'mrpips_data.json')
    with open(mrpips_path, 'w', encoding='utf-8') as f:
        json.dump(mrpips_out, f, ensure_ascii=False, indent=2)

    # Podsumowanie
    total_records = sum(len(v) for v in mrpips_out.values())
    print(f"\n  Zapisano: {mrpips_path}")
    print(f"  Okresy: {len(mrpips_out)}  |  Rekordy: {total_records}")
    print(f"  Rozmiar pliku: {os.path.getsize(mrpips_path) // 1024} KB")

    # Weryfikacja dla najnowszego okresu
    if mrpips_out:
        last_period = sorted(mrpips_out.keys())[-1]
        total_bezr = sum(p.get('bezr_razem', 0) for p in mrpips_out[last_period].values())
        print(f"  Suma bezr. ({last_period}): {total_bezr:,}")

    # ── 4. Wynagrodzenia ─────────────────────────────────────────────────
    print(f"\nWynagrodzenia:")
    if os.path.exists(WYNAGR_FILE):
        print(f"  Czytam: {os.path.basename(WYNAGR_FILE)} ...", end='  ')
        wynagr = extract_wynagrodzenia(WYNAGR_FILE)
        print(f"OK ({len(wynagr.get('powiaty', []))} powiatów, "
              f"{len(wynagr.get('zawody', []))} zawodów)")
    else:
        print(f"  BRAK: {os.path.relpath(WYNAGR_FILE, BASE)}")
        wynagr = {}

    wynagr_path = os.path.join(OUT_DIR, 'wynagr_data.json')
    with open(wynagr_path, 'w', encoding='utf-8') as f:
        json.dump(wynagr, f, ensure_ascii=False, indent=2)
    print(f"  Zapisano: {wynagr_path}")

    # ── 5. Zwolnienia grupowe ─────────────────────────────────────────────
    print(f"\nZwolnienia grupowe:")
    zwolnienia = extract_zwolnienia()
    if zwolnienia:
        print(f"  OK: {len(zwolnienia.get('trend_13m', []))} miesięcy, "
              f"zgl_cur={zwolnienia.get('zgl_cur',0):,}, pkd_top10={len(zwolnienia.get('pkd_top10',[]))}")
    else:
        print("  BRAK danych o zwolnieniach")

    # ── 6. Buduj dashboard_final.json ─────────────────────────────────────
    print(f"\nBudowanie dashboard_final.json:")
    final = build_dashboard_final(mrpips_out, wynagr, zwolnienia)
    final_path = os.path.join(OUT_DIR, 'dashboard_final.json')
    with open(final_path, 'w', encoding='utf-8') as f:
        json.dump(final, f, ensure_ascii=False, indent=2)
    size_kb = os.path.getsize(final_path) // 1024
    print(f"  Zapisano: {final_path}")
    print(f"  Rozmiar:  {size_kb} KB")
    if final.get('bezrobotni'):
        b = final['bezrobotni']
        print(f"  bezr_razem={b['bezr_razem']:,}  wyrej_delta={b['wyrej_delta']:+,}  oferty_delta={b['oferty_delta']:+,}")

    print("\n" + "=" * 65)
    print("  Gotowe! Dane w public/data/ — aplikacja wczyta je automatycznie.")
    print("  Uruchom serwer: npm run dev (port 5173)")
    print("=" * 65)


if __name__ == '__main__':
    main()
