"""
Generator sprawozdań MRPiPS-01
================================
Czyta plik DBF i generuje jeden plik Excel z 42 arkuszami
(po jednym dla każdego powiatu) - identyczny wygląd jak oryginał GUS.

Obsługuje oba formaty:
  - Stary DBF (przed 2026): MP1RRMM14.dbf  np. MP1250514.dbf
  - Nowy DBF (od 2026):     MP1RRMM14.dbf  np. MP1260114.dbf
    lub XLSX z arkuszem dbf: np. 01_2026.xlsx

Użycie:
  py generuj_mrpips.py MP1250514.dbf
  py generuj_mrpips.py 01_2026.xlsx
  py generuj_mrpips.py   (szuka pliku automatycznie w bieżącym folderze)

Wynik: MRPiPS01_POWIAT_RRRR_MM.xlsx
"""

import struct
import os
import sys
import re
import shutil
from copy import copy
from collections import defaultdict

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    print("BŁĄD: Brak biblioteki openpyxl.")
    print("Zainstaluj: py -m pip install openpyxl")
    sys.exit(1)


# ======================================================================
# KONFIGURACJA - NAZWY POWIATÓW I KOLEJNOŚĆ
# ======================================================================

POWIATY = {
    '1401': 'białobrzeski',
    '1402': 'ciechanowski',
    '1403': 'garwoliński',
    '1404': 'gostyniński',
    '1405': 'grodziski',
    '1406': 'grójecki',
    '1407': 'kozienicki',
    '1408': 'legionowski',
    '1409': 'lipski',
    '1410': 'łosicki',
    '1411': 'makowski',
    '1412': 'miński',
    '1413': 'mławski',
    '1414': 'nowodworski',
    '1415': 'ostrołęcki',
    '1416': 'ostrowski',
    '1417': 'otwocki',
    '1418': 'piaseczyński',
    '1419': 'płocki',
    '1420': 'płoński',
    '1421': 'pruszkowski',
    '1422': 'przasnyski',
    '1423': 'przysuski',
    '1424': 'pułtuski',
    '1425': 'radomski',
    '1426': 'siedlecki',
    '1427': 'sierpecki',
    '1428': 'sochaczewski',
    '1429': 'sokołowski',
    '1430': 'szydłowiecki',
    '1432': 'warszawski zachodni',
    '1433': 'węgrowski',
    '1434': 'wołomiński',
    '1435': 'wyszkowski',
    '1436': 'zwoleński',
    '1437': 'żuromiński',
    '1438': 'żyrardowski',
    '1461': 'm. Ostrołęka',
    '1462': 'm. Płock',
    '1463': 'm. Radom',
    '1464': 'm. Siedlce',
    '1465': 'm. Warszawa',
}

NAZWY_MIESIECY = {
    '01': 'styczeń', '02': 'luty', '03': 'marzec',
    '04': 'kwiecień', '05': 'maj', '06': 'czerwiec',
    '07': 'lipiec', '08': 'sierpień', '09': 'wrzesień',
    '10': 'październik', '11': 'listopad', '12': 'grudzień',
}


# ======================================================================
# CZYTANIE DANYCH
# ======================================================================

def read_dbf_binary(filename, encoding='cp1250'):
    """Czyta binarny plik .DBF."""
    with open(filename, 'rb') as f:
        header = f.read(32)
        num_records = struct.unpack('<I', header[4:8])[0]
        header_size = struct.unpack('<H', header[8:10])[0]
        record_size = struct.unpack('<H', header[10:12])[0]
        fields = []
        while True:
            fd = f.read(32)
            if fd[0] == 0x0D:
                break
            name    = fd[:11].replace(b'\x00', b'').decode(encoding)
            ftype   = chr(fd[11])
            length  = fd[16]
            decimal = fd[17]
            fields.append({'name': name, 'type': ftype, 'length': length, 'decimal': decimal})
        f.seek(header_size)
        records = []
        for _ in range(num_records):
            rd = f.read(record_size)
            if not rd or rd[0] == 0x2A:
                continue
            rec = {}
            pos = 1
            for field in fields:
                raw = rd[pos:pos+field['length']].decode(encoding, errors='replace').strip()
                if field['type'] == 'N':
                    try:
                        rec[field['name']] = int(raw) if raw else None
                    except ValueError:
                        rec[field['name']] = None
                else:
                    rec[field['name']] = raw
                pos += field['length']
            records.append(rec)
    return records


def read_xlsx_dbf_sheet(filename):
    """Czyta dane z arkusza 'dbf' w pliku .XLSX (nowy format od 2026)."""
    wb = load_workbook(filename, read_only=True, data_only=True)
    if 'dbf' not in wb.sheetnames:
        raise ValueError(f"Brak arkusza 'dbf' w pliku {filename}")
    ws = wb['dbf']
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        rec = {
            'WGM':    str(row[0]).strip(),
            'M_C':    str(int(row[1])).zfill(2) if row[1] is not None else '',
            'ROK':    str(row[2]).strip() if row[2] else '',
            'DZIAL':  str(row[3]).strip() if row[3] else '',
            'TABELA': str(row[4]).strip() if row[4] else '',
            'NRW':    str(row[5]).strip() if row[5] else '',
        }
        for j in range(1, 23):
            col_idx = 5 + j
            val = row[col_idx] if col_idx < len(row) else None
            try:
                rec[f'R{j}'] = int(val) if val is not None else None
            except (ValueError, TypeError):
                rec[f'R{j}'] = None
        records.append(rec)
    wb.close()
    return records


def wczytaj_dane(filename):
    """Czyta dane z pliku DBF lub XLSX, zwraca (records, rok, miesiac)."""
    ext = os.path.splitext(filename)[1].lower()
    if ext == '.dbf':
        records = read_dbf_binary(filename)
    elif ext in ('.xlsx', '.xlsm'):
        records = read_xlsx_dbf_sheet(filename)
    else:
        raise ValueError(f"Nieobsługiwany format: {ext}")

    if not records:
        raise ValueError("Plik nie zawiera danych")

    sample = next(r for r in records if r.get('DZIAL') != 'C')
    rok     = str(sample.get('ROK', ''))
    miesiac = str(sample.get('M_C', '')).zfill(2)
    return records, rok, miesiac


def buduj_indeks(records):
    """
    Buduje indeks: (WGM, DZIAL, TABELA, NRW) -> rekord.
    Uwaga: stary format (2025) ma inne NRW niż nowy (2026).
    """
    index = {}
    for r in records:
        nrw = str(r.get('NRW', '')).strip()
        key = (str(r.get('WGM','')), str(r.get('DZIAL','')),
               str(r.get('TABELA','')), nrw)
        index[key] = r
    return index


def get_val(index, wgm, dzial, tabela, nrw, kolumna):
    """Pobiera wartość z indeksu. Zwraca None jeśli brak (komórka X w formularzu)."""
    nrw_str = str(nrw).zfill(3) if str(nrw).upper() != 'CZ1' else 'CZ1'
    key = (str(wgm), str(dzial), str(tabela), nrw_str)
    rec = index.get(key)
    if rec is None:
        return None
    return rec.get(kolumna)


# ======================================================================
# WYKRYWANIE WERSJI FORMULARZA
# ======================================================================

def wykryj_wersje(records):
    """
    Wykrywa wersję formularza na podstawie struktury NRW.
    
    Stary format (do 2025):
      D1T1: 20 wierszy (NRW 001-020)
      D2T1: 9 wierszy
      D3T1: 9 wierszy
      Max kolumn: R1-R15
    
    Nowy format (od 2026):
      D1T1: 25 wierszy (NRW 001-025)
      D2T1: 22 wiersze
      D3T1: 7 wierszy
      Max kolumn: R1-R22
    """
    wgm0 = sorted(set(r['WGM'] for r in records if r.get('DZIAL') != 'C'))[0]
    d1t1 = [r for r in records if r['WGM'] == wgm0 and r.get('DZIAL') == '1' and r.get('TABELA') == '1']
    return '2026' if len(d1t1) == 25 else '2025'


# ======================================================================
# MAPOWANIE DBF -> WIERSZE FORMULARZA
# ======================================================================

# Stary format (do 2025): D1T1 NRW 001-020, D1T2 NRW 021-066 itd.
# Nowy format (od 2026):  D1T1 NRW 001-025, D1T2 NRW 026-072 itd.

def buduj_mape_formularza_2025():
    """
    Zwraca listę: (row_excel, dzial, tabela, nrw, kolumna)
    dla arkusza powiatu w formacie 2025.
    Bazuje na odczytanych formułach =dbf!G130 z szablonu.
    """
    # Format wiersza excel -> (dzial, tabela, nrw, [kolumny R])
    # Każda pozycja to jeden wiersz danych w formularzu
    # Oparte na dokładnej analizie szablonu 05_2025.xlsx
    
    mapa = []
    
    # === DZIAŁ 1.1 Struktura bezrobotnych (Row 16-36) ===
    # D1, T1, NRW 001-020, kolumny R1-R8
    d1t1 = [
        (16, '1','1','001'), (17, '1','1','002'), (18, '1','1','003'),
        (19, '1','1','004'), (20, '1','1','000'),  # row 20 = naglowek, brak NRW
        (21, '1','1','005'), (22, '1','1','006'), (23, '1','1','007'),
        (24, '1','1','008'), (25, '1','1','009'), (26, '1','1','010'),
        (27, '1','1','011'), (28, '1','1','012'), (29, '1','1','013'),
        (30, '1','1','014'), (31, '1','1','015'), (32, '1','1','016'),
        (33, '1','1','017'), (34, '1','1','018'), (35, '1','1','019'),
        (36, '1','1','020'),
    ]
    # === DZIAŁ 1.2 Bilans bezrobotnych (Row 44-89) ===
    d1t2 = [
        (44, '1','2','021'), (45, '1','2','022'), (46, '1','2','023'),
        (47, '1','2','024'), (48, '1','2','025'), (49, '1','2','026'),
        (50, '1','2','027'), (51, '1','2','028'), (52, '1','2','029'),
        (53, '1','2','030'), (54, '1','2','031'), (55, '1','2','032'),
        (56, '1','2','033'), (57, '1','2','034'), (58, '1','2','035'),
        (59, '1','2','036'), (60, '1','2','037'), (61, '1','2','038'),
        (62, '1','2','039'), (63, '1','2','040'), (64, '1','2','041'),
        (65, '1','2','042'), (66, '1','2','043'), (67, '1','2','044'),
        (68, '1','2','045'), (69, '1','2','046'), (70, '1','2','047'),
        (71, '1','2','048'), (72, '1','2','049'), (73, '1','2','050'),
        (74, '1','2','051'), (75, '1','2','052'), (76, '1','2','053'),
        (77, '1','2','054'), (78, '1','2','055'), (79, '1','2','056'),
        (80, '1','2','057'), (81, '1','2','058'), (82, '1','2','059'),
        (83, '1','2','060'), (84, '1','2','061'), (85, '1','2','062'),
        (86, '1','2','063'), (87, '1','2','064'), (88, '1','2','065'),
        (89, '1','2','066'),
    ]
    # === DZIAŁ 1.3 Aktywne formy (Row 99-113) ===
    d1t3 = [
        (99,  '1','3','067'), (100, '1','3','068'), (101, '1','3','069'),
        (102, '1','3','070'), (103, '1','3','071'), (104, '1','3','072'),
        (105, '1','3','073'), (106, '1','3','074'), (107, '1','3','075'),
        (108, '1','3','076'), (109, '1','3','077'), (110, '1','3','078'),
        (111, '1','3','079'), (112, '1','3','080'), (113, '1','3','081'),
    ]
    # === DZIAŁ 2 Poszukujący pracy (Row 120-128) ===
    d2 = [
        (120, '2','1','001'), (121, '2','1','002'), (122, '2','1','003'),
        (123, '2','1','004'), (124, '2','1','005'), (125, '2','1','006'),
        (126, '2','1','007'), (127, '2','1','008'), (128, '2','1','009'),
    ]
    # === DZIAŁ 3 Wolne miejsca pracy (Row 136-144) ===
    d3 = [
        (136, '3','1','001'), (137, '3','1','002'), (138, '3','1','003'),
        (139, '3','1','004'), (140, '3','1','005'), (141, '3','1','006'),
        (142, '3','1','007'), (143, '3','1','008'), (144, '3','1','009'),
    ]
    # === DZIAŁ 4 Zwolnienia grupowe (Row 151-153) ===
    d4 = [
        (151, '4','1','001'), (152, '4','1','002'), (153, '4','1','003'),
    ]
    # === DZIAŁ 5 Struktura wg cech (Row 162-186) ===
    d5 = [
        (162, '5','1','001'), (163, '5','1','002'), (164, '5','1','003'),
        (165, '5','1','004'), (166, '5','1','005'), (167, '5','1','006'),
        (168, '5','1','007'), (169, '5','1','008'), (170, '5','1','009'),
        (171, '5','1','010'), (172, '5','1','011'), (173, '5','1','012'),
        (174, '5','1','013'), (175, '5','1','014'), (176, '5','1','015'),
        (177, '5','1','016'), (178, '5','1','017'), (179, '5','1','018'),
        (180, '5','1','019'), (181, '5','1','020'), (182, '5','1','021'),
        (183, '5','1','022'), (184, '5','1','023'), (185, '5','1','024'),
        (186, '5','1','025'),
    ]
    
    return d1t1 + d1t2 + d1t3 + d2 + d3 + d4 + d5


def buduj_mape_formularza_2026():
    """Mapa dla nowego formularza 2026."""
    d1t1 = [(16+i, '1','1',f'{i+1:03}') for i in range(25)]
    d1t2 = [(49+i, '1','2',f'{26+i:03}') for i in range(47)]  # NRW 026-072
    # ... (strukturę zbudujemy dynamicznie z szablonu)
    return d1t1 + d1t2



# ======================================================================
# WYSZUKIWANIE SZABLONU
# ======================================================================

def _znajdz_szablon(folder, rok):
    """
    Szuka szablonu GUS w podanym folderze.

    Kolejność priorytetów:
      1. szablon_RRRR.xlsx         - dokładnie pasujący rok (np. szablon_2026.xlsx)
      2. szablon_RRRR.xlsx         - najbliższy wcześniejszy rok
      3. szablon.xlsx              - ogólny szablon bez roku
      4. Dowolny .xlsx z arkuszami powiatów (10+ arkuszy cyfrowych)

    Zalecane nazewnictwo:
      szablon_2025.xlsx  <- dla danych z lat do 2025 włącznie
      szablon_2026.xlsx  <- dla danych od 2026
    """
    rok_int = int(rok) if rok else 0

    # Szukaj pliku szablon_RRRR.xlsx - wybierz najlepiej pasujący rok
    szablon_lata = {}
    for f in os.listdir(folder):
        m = re.match(r'^szablon_(\d{4})\.xlsx$', f, re.IGNORECASE)
        if m:
            szablon_lata[int(m.group(1))] = os.path.join(folder, f)

    if szablon_lata:
        # Znajdź największy rok <= rok danych
        pasujace = [r for r in szablon_lata if r <= rok_int]
        if pasujace:
            wybrany_rok = max(pasujace)
            return szablon_lata[wybrany_rok]
        # Jeśli nie ma starszego, weź najmniejszy dostępny
        return szablon_lata[min(szablon_lata)]

    # Szukaj pliku szablon.xlsx
    ogolny = os.path.join(folder, 'szablon.xlsx')
    if os.path.exists(ogolny):
        return ogolny

    # Fallback: dowolny .xlsx z minimum 10 arkuszami cyfrowymi (powiatami)
    kandydaci = []
    for f in sorted(os.listdir(folder)):
        if not f.lower().endswith('.xlsx'):
            continue
        full = os.path.join(folder, f)
        try:
            wb_test = load_workbook(full, read_only=True)
            sheets = wb_test.sheetnames
            wb_test.close()
            powiat_sheets = [s for s in sheets if s.isdigit() and len(s) <= 2]
            if len(powiat_sheets) >= 10:
                kandydaci.append(full)
        except Exception:
            pass

    if kandydaci:
        # Preferuj szablon z rokiem jak najbliższym danym
        def rok_z_nazwy(path):
            m = re.search(r'(\d{4})', os.path.basename(path))
            return int(m.group(1)) if m else 0

        kandydaci.sort(key=lambda p: abs(rok_z_nazwy(p) - rok_int))
        return kandydaci[0]

    raise FileNotFoundError(
        f"Nie znaleziono szablonu GUS dla roku {rok}!\n"
        f"\nUmieść w folderze jeden z poniższych plików:\n"
        f"  szablon_{rok}.xlsx   <- zalecane (dokładny rok)\n"
        f"  szablon_2025.xlsx    <- dla danych do 2025\n"
        f"  szablon_2026.xlsx    <- dla danych od 2026\n"
        f"  szablon.xlsx         <- ogólny\n"
        f"\nSzablon to plik XLSX od GUS z formularzem MRPiPS-01\n"
        f"zawierający arkusze powiatów (01, 02, 03 ... 65)."
    )


# ======================================================================
# GENERATOR GŁÓWNY
# ======================================================================

def znajdz_formuly_w_szablonie(ws_template):
    """
    Skanuje arkusz szablonu i wyciąga mapowanie:
    (row, col) -> numer_wiersza_dbf (z formuły =dbf!G130)
    Zwraca dict: (row, col) -> dbf_row_num
    """
    mapa = {}
    pattern = re.compile(r'=dbf!([A-Z]+)(\d+)', re.IGNORECASE)
    for row in ws_template.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                m = pattern.match(cell.value.strip())
                if m:
                    col_letter = m.group(1).upper()
                    dbf_row    = int(m.group(2))
                    mapa[(cell.row, cell.column)] = (col_letter, dbf_row)
    return mapa


def kopiuj_styl_komorki(src, dst):
    """Kopiuje styl (font, fill, border, alignment) z src do dst."""
    if src.has_style:
        dst.font      = copy(src.font)
        dst.fill      = copy(src.fill)
        dst.border    = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format


def generuj_sprawozdania(plik_dbf, plik_szablon=None, plik_wyjsciowy=None):
    """
    Główna funkcja generatora.
    
    Args:
        plik_dbf:       ścieżka do pliku DBF (lub XLSX z arkuszem dbf)
        plik_szablon:   ścieżka do pliku XLSX z szablonem GUS (opcjonalnie)
        plik_wyjsciowy: ścieżka do pliku wyjściowego (opcjonalnie)
    """
    
    # --- Wczytaj dane ---
    print(f"\n[1/4] Wczytuję dane z: {os.path.basename(plik_dbf)}")
    records, rok, miesiac = wczytaj_dane(plik_dbf)
    index = buduj_indeks(records)
    wersja = wykryj_wersje(records)
    powiaty_w_danych = sorted(set(r['WGM'] for r in records if r.get('DZIAL') != 'C'))
    
    nazwa_miesiaca = NAZWY_MIESIECY.get(miesiac, miesiac)
    print(f"    Wersja formularza: {wersja}")
    print(f"    Okres: {rok}-{miesiac} ({nazwa_miesiaca} {rok})")
    print(f"    Powiatów: {len(powiaty_w_danych)}")
    
    # --- Znajdź szablon ---
    if plik_szablon is None:
        folder = os.path.dirname(os.path.abspath(plik_dbf))
        plik_szablon = _znajdz_szablon(folder, rok)
        print(f"    Szablon: {os.path.basename(plik_szablon)}")
    else:
        print(f"    Szablon: {os.path.basename(plik_szablon)}")
    
    # --- Wczytaj szablon ---
    print(f"\n[2/4] Wczytuję szablon formularza...")
    wb_szablon = load_workbook(plik_szablon)
    
    # Znajdź arkusze powiatów w szablonie
    arkusze_powiatow = sorted(
        [s for s in wb_szablon.sheetnames if s.isdigit() and len(s) <= 2],
        key=int
    )
    print(f"    Znaleziono {len(arkusze_powiatow)} arkuszy powiatów w szablonie")
    
    # Weź pierwszy arkusz powiatu jako wzorzec
    wzorzec_sheet = arkusze_powiatow[0]
    ws_wzorzec = wb_szablon[wzorzec_sheet]
    
    # Wyciągnij mapę formuł: (row,col) -> (litera_kolumny_dbf, numer_wiersza_dbf)
    formuly = znajdz_formuly_w_szablonie(ws_wzorzec)
    print(f"    Znaleziono {len(formuly)} formuł =dbf!... w szablonie")
    
    # Sprawdź ile wierszy zajmuje jeden powiat w arkuszu dbf szablonu
    # (potrzebne do obliczenia offsetu dla każdego powiatu)
    ws_dbf_szablon = wb_szablon['dbf']
    powiaty_w_szablonie = []
    prev_wgm = None
    start_row = {}
    for i, row_data in enumerate(ws_dbf_szablon.iter_rows(min_row=2, values_only=True), 2):
        wgm = str(row_data[0]) if row_data[0] else None
        if wgm and wgm != prev_wgm:
            start_row[wgm] = i
            if prev_wgm:
                powiaty_w_szablonie.append(prev_wgm)
            prev_wgm = wgm
    if prev_wgm:
        powiaty_w_szablonie.append(prev_wgm)
    
    # Wierszy na powiat = offset między kolejnymi powiatami
    if len(powiaty_w_szablonie) >= 2:
        WIERSZY_NA_POWIAT = start_row[powiaty_w_szablonie[1]] - start_row[powiaty_w_szablonie[0]]
    else:
        WIERSZY_NA_POWIAT = sum(1 for r in records if r['WGM'] == powiaty_w_danych[0])
    
    # Oblicz offset wzorca (numer pierwszego wiersza powiatu we wzorcu)
    # Wzorzec to arkusz_powiatow[0], który wskazuje na pewien wiersz w dbf
    # Odczytaj N1 z wzorca aby wiedzieć jaki to wiersz
    n1_val = ws_wzorzec['N1'].value
    if n1_val and 'dbf!A' in str(n1_val):
        wzorzec_dbf_row = int(re.search(r'A(\d+)', str(n1_val)).group(1))
    else:
        wzorzec_dbf_row = 2  # domyślnie pierwszy powiat
    
    # WGM wzorca (pierwszy powiat w szablonie)
    wzorzec_wgm = str(ws_dbf_szablon.cell(wzorzec_dbf_row, 1).value)
    
    print(f"    Wzorzec: arkusz '{wzorzec_sheet}' -> WGM={wzorzec_wgm}, dbf row={wzorzec_dbf_row}")
    print(f"    Wierszy na powiat w szablonie: {WIERSZY_NA_POWIAT}")
    
    # --- Mapowanie: kolumna litery -> numer R ---
    # W arkuszu dbf szablonu kolumny to: A=WGM, B=M_C, C=ROK, D=DZIAL, E=TABELA, F=NRW,
    # G=R1, H=R2, I=R3, J=R4, K=R5, L=R6, M=R7, N=R8 itd.
    def col_letter_to_R(letter):
        col_num = column_index_from_string(letter)
        r_num = col_num - 6  # G=7 -> R1, H=8 -> R2 itd.
        return f'R{r_num}' if r_num >= 1 else None
    
    def col_letter_to_field(letter):
        col_num = column_index_from_string(letter)
        if col_num == 1: return 'WGM'
        if col_num == 2: return 'M_C'
        if col_num == 3: return 'ROK'
        if col_num == 4: return 'DZIAL'
        if col_num == 5: return 'TABELA'
        if col_num == 6: return 'NRW'
        r_num = col_num - 6
        return f'R{r_num}' if r_num >= 1 else None
    
    # --- Buduj mapę: nr_wiersza_w_wzorcu -> rekord_dbf ---
    # Odczytaj wszystkie wiersze dbf-szablonu dla wzorca i zapisz strukturę
    dbf_wiersze_wzorca = {}  # dbf_row -> {field: value}
    for i in range(WIERSZY_NA_POWIAT):
        row_num = wzorzec_dbf_row + i
        row_data = [ws_dbf_szablon.cell(row_num, c).value for c in range(1, 28)]
        dbf_wiersze_wzorca[row_num] = row_data
    
    # --- Twórz nowy plik wyjściowy ---
    if plik_wyjsciowy is None:
        plik_wyjsciowy = os.path.join(os.getcwd(), f"MRPiPS01_POWIATY_{rok}_{miesiac}.xlsx")
    
    print(f"\n[3/4] Generuję plik wyjściowy...")
    print(f"    Plik: {os.path.basename(plik_wyjsciowy)}")
    
    # Skopiuj cały plik szablonu jako bazę
    shutil.copy2(plik_szablon, plik_wyjsciowy)
    wb_out = load_workbook(plik_wyjsciowy)
    
    # Zaktualizuj arkusz dbf danymi z nowego DBF
    ws_dbf_out = wb_out['dbf']
    
    # Wyczyść stare dane (zostaw nagłówek)
    for row in ws_dbf_out.iter_rows(min_row=2):
        for cell in row:
            cell.value = None
    
    # Wpisz nowe dane
    for i, wgm in enumerate(powiaty_w_danych):
        base_row = 2 + i * WIERSZY_NA_POWIAT
        rekordy_powiatu = [r for r in records if r['WGM'] == wgm]
        # Posortuj tak samo jak oryginał (DZIAL, TABELA, NRW)
        rekordy_powiatu.sort(key=lambda r: (r.get('DZIAL',''), r.get('TABELA',''), r.get('NRW','')))
        
        for j, rec in enumerate(rekordy_powiatu):
            row_num = base_row + j
            ws_dbf_out.cell(row_num, 1, rec.get('WGM', ''))
            ws_dbf_out.cell(row_num, 2, rec.get('M_C', ''))
            ws_dbf_out.cell(row_num, 3, rec.get('ROK', ''))
            ws_dbf_out.cell(row_num, 4, rec.get('DZIAL', ''))
            ws_dbf_out.cell(row_num, 5, rec.get('TABELA', ''))
            ws_dbf_out.cell(row_num, 6, rec.get('NRW', ''))
            for r_num in range(1, 23):
                val = rec.get(f'R{r_num}')
                ws_dbf_out.cell(row_num, 6 + r_num, val)
    
    print(f"    Zaktualizowano arkusz 'dbf' ({len(powiaty_w_danych) * WIERSZY_NA_POWIAT} wierszy)")
    
    # Zaktualizuj Arkusz2 (miesiące)
    if 'Arkusz2' in wb_out.sheetnames:
        ws_ark2 = wb_out['Arkusz2']
        # Zaktualizuj rok w nazwach miesięcy
        for row in ws_ark2.iter_rows(min_row=1, max_row=12):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and 'miesiąc' in cell.value:
                    # Podmień rok np. "za miesiąc styczeń 2025" -> "za miesiąc styczeń 2026"
                    stary_rok = re.search(r'\d{4}', cell.value)
                    if stary_rok and stary_rok.group() != rok:
                        cell.value = cell.value.replace(stary_rok.group(), rok)
    
    # Zaktualizuj komórki B2 w dbf (miesiąc) - pierwszy wiersz może mieć liczbę
    # sprawdź czy M_C jest poprawnie zapisane
    
    # Zapisz plik
    wb_out.save(plik_wyjsciowy)
    print(f"    Zapisano plik (formuły w arkuszach powiatów zaktualizują się automatycznie)")
    
    print(f"\n[4/4] Gotowe!")
    print(f"    Plik: {plik_wyjsciowy}")
    print(f"    Arkusze powiatów: {len(arkusze_powiatow)} (formuły =dbf!... pobiorą nowe dane)")
    print(f"\n    Otwórz plik w Excel - każdy arkusz powiatu wyświetli")
    print(f"    dane za {nazwa_miesiaca} {rok} r.")
    
    return plik_wyjsciowy


# ======================================================================
# URUCHOMIENIE
# ======================================================================

def znajdz_plik_dbf(folder='.'):
    """Automatycznie znajduje plik DBF lub XLSX z danymi w podanym folderze."""
    kandidaci_dbf  = [f for f in os.listdir(folder) if f.upper().startswith('MP1') and f.lower().endswith('.dbf')]
    kandidaci_xlsx = [f for f in os.listdir(folder) if re.match(r'^\d{2}_\d{4}\.xlsx$', f)]
    
    if kandidaci_dbf:
        return os.path.join(folder, sorted(kandidaci_dbf)[-1])  # najnowszy
    if kandidaci_xlsx:
        return os.path.join(folder, sorted(kandidaci_xlsx)[-1])
    return None


if __name__ == '__main__':
    folder = os.path.dirname(os.path.abspath(__file__))
    
    # Argument z linii poleceń lub automatyczne wyszukiwanie
    if len(sys.argv) > 1:
        plik_dbf = sys.argv[1]
        if not os.path.isabs(plik_dbf):
            plik_dbf = os.path.join(folder, plik_dbf)
        if not os.path.exists(plik_dbf):
            print(f"BŁĄD: Nie znaleziono pliku: {plik_dbf}")
            sys.exit(1)
    else:
        plik_dbf = znajdz_plik_dbf(folder)
        if not plik_dbf:
            print("BŁĄD: Nie znaleziono pliku DBF ani XLSX z danymi.")
            print("Umieść plik MP1RRMM14.dbf lub MM_RRRR.xlsx w tym samym folderze.")
            sys.exit(1)
        print(f"Znaleziono automatycznie: {os.path.basename(plik_dbf)}")
    
    try:
        wynik = generuj_sprawozdania(plik_dbf)
        print(f"\nSUKCES! Otwórz: {wynik}")
    except FileNotFoundError as e:
        print(f"\nBŁĄD: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"\nNieoczekiwany błąd: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
