"""
MRPiPS-01 Parser - czytanie danych ze sprawozdan o rynku pracy
Autor: na podstawie oficjalnej dokumentacji GUS/MRPiPS

Obsluguje OBA formaty:
  - Stary format (przed 2026): plik .DBF  (np. MP1260114.dbf)
  - Nowy format (od 2026):     plik .XLSX (np. 01_2026.xlsx, arkusz 'dbf')

Struktura danych jest IDENTYCZNA w obu formatach:
  WGM, M_C, ROK, DZIAL, TABELA, NRW, R1..R22

=======================================================================
OFICJALNA STRUKTURA DANYCH (wg dokumentacji GUS)
=======================================================================

Nazwa pliku DBF: MP1MMWWXX
  Schemat praktyczny: MP1 + rok(2 cyfry) + MM + WW
  Przyklad: MP1260114 -> rok=2026, miesiac=01, woj=14 (mazowieckie)

Pola:
  Nr  Nazwa   Typ     Dl  Opis
  1   WGM     Znak    4   Symbol powiatu: 2 cyfry woj (02-32) + 2 cyfry pow (01-79)
  2   M_C     Znak    2   Miesiac 01-12
  3   ROK     Znak    4   Rok
  4   DZIAL   Znak    1   Dzial: 1-5 plus C (czas wypelnienia)
  5   TABELA  Znak    1   Tabela w dziale: 1-3
  6   NRW     Znak    3   Numer wiersza formularza (np. '001', 'CZ1')
  7   R1      Num     6   }
  ..  ..      ..      ..  } Rubryki R1-R22: wartosci z kolumn formularza
  21  R22     Num     6   }

Specjalny DZIAL C:
  NRW='CZ1', R1=czas przygotowania danych (min), R2=czas wypelnienia (min)

=======================================================================
MAPOWANIE: DBF -> ARKUSZE XLSX (tablice kontrolne)
=======================================================================

  DZIAL 1, TABELA 1  : wiersze NRW 001-025  (Struktura bezrobotnych)
  DZIAL 1, TABELA 2  : wiersze NRW 026-072  (Bilans bezrobotnych)
  DZIAL 1, TABELA 3  : wiersze NRW 073-100  (wg czasu pozostawania)
  DZIAL 2, TABELA 1  : wiersze NRW 001-022  (Poszukujacy pracy)
  DZIAL 3, TABELA 1  : wiersze NRW 001-007  (Wolne miejsca pracy)
  DZIAL 4, TABELA 1  : wiersze NRW 001-003  (Zwolnienia grupowe)
  DZIAL 5, TABELA 1  : wiersze NRW 001-025  (Struktura wg cech)
  DZIAL C, TABELA 1  : wiersz  NRW CZ1      (Czas wypelnienia)

  Lacznie: 158 wierszy na powiat, 42 powiaty = 6636 rekordow

=======================================================================
ROZNICE MIEDZY FORMATAMI
=======================================================================

  Cecha                | Stary format (.DBF)        | Nowy format (.XLSX od 2026)
  ---------------------+----------------------------+---------------------------
  Plik                 | MP1260114.dbf              | 01_2026.xlsx
  Schemat nazwy        | MP1 + RR + MM + WW         | MM_RRRR.xlsx
  Dane                 | plik binarny .dbf          | arkusz 'dbf' w pliku xlsx
  Struktura danych     | identyczna                 | identyczna
  Dodatkowe arkusze    | brak                       | 42 powiaty + woj + regiony
  Arkusze powiatow     | brak                       | formuly =dbf!G2 itd.
  Kodowanie            | cp1250 (Windows)           | UTF-8 (Excel)
  Wymagane biblioteki  | tylko struct (wbudowany)   | openpyxl

=======================================================================
SZYBKI KLUCZ: CO I GDZIE
=======================================================================

  Pytanie                                 | DZ TAB  NRW   R
  ----------------------------------------|------------------
  Zarejestrowali sie (ogolem)?            |  1   1  001   R1
  Zarejestrowali sie (kobiety)?           |  1   1  001   R2
  Stan na koniec miesiaca (ogolem)?       |  1   1  001   R5
  Stan na koniec miesiaca (kobiety)?      |  1   1  001   R6
  Z prawem do zasilku (koniec mies.)?     |  1   1  001   R7
  Wylaczeni z ewidencji (ogolem)?         |  1   2  034   R1
  Podjeli prace (ogolem)?                 |  1   2  035   R1
  Dlugotrwale bezrobotni (koniec mies.)?  |  1   1  021   R5
  Niepelnosprawni (koniec mies.)?         |  1   1  023   R5
  Poszukujacy pracy (koniec mies.)?       |  2   1  004   R3
  Wolne miejsca pracy (ogolem)?           |  3   1  001   R1
  Zgloszone zwolnienia grupowe?           |  4   1  001   R2
  Czas wypelnienia formularza?            |  C   1  CZ1   R1/R2
"""

import struct
import os
from collections import defaultdict


# ======================================================================
# OPISY WIERSZY FORMULARZA
# ======================================================================

OPIS_WIERSZY = {
    # DZIAL 1, TABELA 1: Struktura bezrobotnych (NRW 001-025)
    ('1', '1'): {
        '001': 'Ogolem (w.02+04)',
        '002': 'Poprzednio pracujace',
        '003': '  w tym zwolnione z przyczyn dot. zakladu pracy',
        '004': 'Dotychczas niepracujace',
        '005': 'Zamieszkali na wsi',
        '006': '  w tym posiadajacy gospodarstwo rolne',
        '007': 'Bez doswiadczenia zawodowego',
        '008': 'Absolwenci (do 12 mies. od ukonczenia nauki)',
        '009': 'Kobiety niepodejmujace zatrudn. po urodzeniu dziecka',
        '010': 'Majacy co najmniej 1 dziecko do 18 lat',
        '011': '  - 1 dziecko',
        '012': '  - 2 dzieci',
        '013': '  - 3 i wiecej dzieci',
        '014': 'Opiekunowie osoby niepelnosprawnej',
        '015': 'Cudzoziemcy',
        '016': '  w tym z krajow EOG oraz Szwajcarii',
        '017': 'Z pierwszenstwem do form pomocy',
        '018': '  - do 30 roku zycia',
        '019': '    - w tym do 25 roku zycia',
        '020': '  - powyzej 50 roku zycia',
        '021': '  - dlugotrwale bezrobotni',
        '022': '  - bez kwalifikacji zawodowych',
        '023': '  - niepelnosprawni',
        '024': '  - posiadajace Karte Duzej Rodziny',
        '025': '  - samotnie wychowujacy co najmniej jedno dziecko',
    },
    # DZIAL 1, TABELA 2: Bilans bezrobotnych (NRW 026-072)
    ('1', '2'): {
        '026': 'Zarejestrowani w miesiacu (w.27+28)',
        '027': '  po raz pierwszy',
        '028': '  po raz kolejny',
        '029': '  po pracach interwencyjnych',
        '030': '  po robotach publicznych',
        '031': '  po stazu',
        '032': '  po szkoleniu',
        '033': '  po pracach spolecznie uzytecznych',
        '034': 'Wylaczeni z ewidencji',
        '035': '  podjeli prace (ogolem)',
        '036': '    - niesubsydiowana',
        '037': '    w tym: dzialalnosc gospodarcza',
        '038': '    w tym: praca sezonowa',
        '039': '    - subsydiowana (w.40-53)',
        '040': '      prace interwencyjne',
        '041': '      roboty publiczne',
        '042': '      podjecie dzialalnosci gospodarczej',
        '043': '      w tym: bon na zasiedlenie',
        '044': '      refundacja kosztow zatrudnienia',
        '045': '      pozyczka na stanowisko pracy',
        '046': '      praca poza miejscem zamieszkania',
        '047': '      swiadczenie aktywizacyjne',
        '048': '      grant na teleprace',
        '049': '      zwrot kosztow przejazdu',
        '050': '      refundacja skladek ZUS',
        '051': '      dofinansowanie wynagrodzenia 50+',
        '052': '      zalozenie/przystapiecie do spoldzielni socjalnej',
        '053': '      inne',
        '054': '  rozpoczeli szkolenie',
        '055': '    w tym: bon szkoleniowy',
        '056': '  rozpoczeli staz',
        '057': '  prace spolecznie uzyteczne',
        '058': '  skierowanie do dzialan reintegracji spoleczno-zaw.',
        '059': '  skierowanie do agencji zatrudnienia',
        '060': '  niepodjecie/przerwanie z wlasnej winy',
        '061': '  skierowanie do agencji w ramach zlecenia dzialan',
        '062': '  odmowa udzialu w IPD',
        '063': '  brak kontaktu z PUP co najmniej 90 dni',
        '064': '  dobrowolna rezygnacja ze statusu bezrobotnego',
        '065': '  podjecie nauki',
        '066': '  osiagniecie wieku emerytalnego',
        '067': '  nabycie praw emerytalnych/rentowych',
        '068': '  swiadczenie przedemerytalne',
        '069': '  inne przyczyny',
        '070': 'Bezrobotni zmieniajacy miejsce zamieszkania',
        '071': 'Stan wg konca miesiaca (ogolem)',
        '072': '  w tym zarejestrowani po raz pierwszy',
    },
    # DZIAL 1, TABELA 3: wg czasu pozostawania bez pracy (NRW 073-100)
    ('1', '3'): {
        '073': 'do 1 miesiaca',
        '074': '1-3 miesiace',
        '075': '3-6 miesiecy',
        '076': '6-12 miesiecy',
        '077': '12-24 miesiace',
        '078': 'pow. 24 miesiecy',
        '079': 'do 1 mies. - kobiety',
        '080': '1-3 mies. - kobiety',
        '081': '3-6 mies. - kobiety',
        '082': '6-12 mies. - kobiety',
        '083': '12-24 mies. - kobiety',
        '084': 'pow. 24 mies. - kobiety',
        '085': 'dlugotrwale bezrobotni (pow. 12 mies.)',
        '086': '  - kobiety',
        '087': 'zarejestrowani do 1 mies.',
        '088': '  - kobiety',
        '089': 'stan koniec mies. do 1 mies.',
        '090': 'stan koniec mies. 1-3 mies.',
        '091': 'stan koniec mies. 3-6 mies.',
        '092': 'stan koniec mies. 6-12 mies.',
        '093': 'stan koniec mies. 12-24 mies.',
        '094': 'stan koniec mies. pow. 24 mies.',
        '095': 'kobiety stan do 1 mies.',
        '096': 'kobiety stan 1-3 mies.',
        '097': 'kobiety stan 3-6 mies.',
        '098': 'kobiety stan 6-12 mies.',
        '099': 'kobiety stan 12-24 mies.',
        '100': 'kobiety stan pow. 24 mies.',
    },
    # DZIAL 2, TABELA 1: Poszukujacy pracy (NRW 001-022)
    ('2', '1'): {
        '001': 'Nabycie uprawnienia do dodatku aktywiz. - dzialalnosc gospod.',
        '002': 'Nabycie uprawnienia do dodatku aktywiz. - zatrudnienie',
        '003': 'Polacy z prawem do zasilku transferowego',
        '004': 'Poszukujacy pracy',
        '005': '  w tym niepelnosprawni niepozostajacy w zatrudnieniu',
        '006': '  opiekunowie osoby niepelnosprawnej',
        '007': '    w tym niezatrudnieni',
        '008': 'Poszukujacy - podjeli prace niesubsydiowana',
        '009': 'Poszukujacy - podjeli prace subsydiowana',
        '010': '  prace interwencyjne',
        '011': '  roboty publiczne',
        '012': '  dzialalnosc gospodarcza',
        '013': '  refundacja kosztow zatrudnienia',
        '014': '  pozyczka na stanowisko pracy',
        '015': '  swiadczenie aktywizacyjne',
        '016': '  grant na teleprace',
        '017': '  zwrot kosztow przejazdu',
        '018': '  refundacja skladek ZUS',
        '019': '  dofinansowanie wynagrodzenia',
        '020': '  spoldzielnia socjalna',
        '021': '  szkolenia',
        '022': '  staze',
    },
    # DZIAL 3, TABELA 1: Wolne miejsca pracy (NRW 001-007)
    ('3', '1'): {
        '001': 'Wolne miejsca pracy - ogolem',
        '002': '  zatrudnienie lub inna praca zarobkowa',
        '003': '  miejsca aktywizacji zawodowej',
        '004': '    - staze',
        '005': '    - prace spolecznie uzyteczne',
        '006': '  dla niepelnosprawnych',
        '007': '  dla absolwentow',
    },
    # DZIAL 4, TABELA 1: Zwolnienia grupowe (NRW 001-003)
    ('4', '1'): {
        '001': 'Zgloszone zwolnienia grupowe',
        '002': 'Zwolnienia grupowe (faktyczne)',
        '003': 'Zwolnienia monitorowane',
    },
    # DZIAL 5, TABELA 1: Struktura wg cech (NRW 001-025)
    ('5', '1'): {
        '001': 'Ogolem',
        '002': 'Czas pozostawania bez pracy: do 1 mies.',
        '003': '  1-3 mies.',
        '004': '  3-6 mies.',
        '005': '  6-12 mies.',
        '006': '  12-24 mies.',
        '007': '  pow. 24 mies.',
        '008': 'Wiek: 18-24',
        '009': '  25-34',
        '010': '  35-44',
        '011': '  45-54',
        '012': '  55-59',
        '013': '  60 lat i wiecej',
        '014': 'Wyksztalcenie: wyzsze',
        '015': '  policealne i srednie zawodowe/branzowe',
        '016': '  srednie ogolnoksztalcace',
        '017': '  zasadnicze zawodowe/branzowe',
        '018': '  gimnazjalne/podstawowe i ponizej',
        '019': 'Staz pracy: do 1 roku',
        '020': '  1-5 lat',
        '021': '  5-10 lat',
        '022': '  10-20 lat',
        '023': '  20-30 lat',
        '024': '  30 lat i wiecej',
        '025': '  bez stazu',
    },
}


# ======================================================================
# CZYTANIE DANYCH
# ======================================================================

def _read_dbf_file(filename: str, encoding: str = 'cp1250') -> list:
    """
    Wczytuje stary format .DBF (przed 2026).
    Nie wymaga zewnetrznych bibliotek - uzywa tylko modulu struct.
    """
    with open(filename, 'rb') as f:
        header = f.read(32)
        num_records  = struct.unpack('<I', header[4:8])[0]
        header_size  = struct.unpack('<H', header[8:10])[0]
        record_size  = struct.unpack('<H', header[10:12])[0]

        fields = []
        while True:
            fd = f.read(32)
            if fd[0] == 0x0D:
                break
            name    = fd[:11].replace(b'\x00', b'').decode(encoding)
            ftype   = chr(fd[11])
            length  = fd[16]
            decimal = fd[17]
            fields.append({'name': name, 'type': ftype, 'length': length, 'decimal': decimal})

        f.seek(header_size)
        records = []
        for _ in range(num_records):
            rd = f.read(record_size)
            if not rd or rd[0] == 0x2A:   # rekord usuniety
                continue
            rec = {}
            pos = 1
            for field in fields:
                raw = rd[pos:pos + field['length']].decode(encoding, errors='replace').strip()
                if field['type'] == 'N':
                    try:
                        rec[field['name']] = int(raw) if raw else 0
                    except ValueError:
                        rec[field['name']] = 0
                else:
                    rec[field['name']] = raw
                pos += field['length']
            records.append(rec)

    return records


def _read_xlsx_file(filename: str) -> list:
    """
    Wczytuje nowy format .XLSX (od 2026).
    Dane sa w arkuszu 'dbf' - identyczna struktura jak plik .dbf.
    Wymaga: openpyxl (pip install openpyxl)
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "Brak biblioteki openpyxl. Zainstaluj: pip install openpyxl"
        )

    wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)

    if 'dbf' not in wb.sheetnames:
        raise ValueError(
            f"Plik {filename} nie zawiera arkusza 'dbf'.\n"
            f"Dostepne arkusze: {wb.sheetnames}"
        )

    ws = wb['dbf']
    records = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rec = {
            'WGM':    str(row[0]).strip(),
            'M_C':    str(int(row[1])).zfill(2) if row[1] is not None else '',
            'ROK':    str(row[2]).strip() if row[2] else '',
            'DZIAL':  str(row[3]).strip() if row[3] else '',
            'TABELA': str(row[4]).strip() if row[4] else '',
            'NRW':    str(row[5]).strip() if row[5] else '',
        }
        for j in range(1, 23):
            col_idx = 5 + j   # kolumny 6..27 (0-indeksowane: 6..27)
            val = row[col_idx] if col_idx < len(row) else None
            try:
                rec[f'R{j}'] = int(val) if val is not None else 0
            except (ValueError, TypeError):
                rec[f'R{j}'] = 0

        records.append(rec)

    wb.close()
    return records


def wczytaj_mrpips(filename: str, encoding: str = 'cp1250') -> list:
    """
    Uniwersalna funkcja wczytujaca dane MRPiPS-01.
    Automatycznie wykrywa format na podstawie rozszerzenia pliku.

    Args:
        filename: sciezka do pliku .dbf lub .xlsx
        encoding: kodowanie dla plikow .dbf (domyslnie cp1250)

    Returns:
        Lista rekordow (slownikow): WGM, M_C, ROK, DZIAL, TABELA, NRW, R1..R22

    Przyklad:
        rekordy = wczytaj_mrpips('MP1260114.dbf')    # stary format
        rekordy = wczytaj_mrpips('01_2026.xlsx')     # nowy format
    """
    ext = os.path.splitext(filename)[1].lower()
    if ext == '.dbf':
        return _read_dbf_file(filename, encoding)
    elif ext in ('.xlsx', '.xlsm'):
        return _read_xlsx_file(filename)
    else:
        raise ValueError(
            f"Nieobslugiwany format: {ext}. Oczekiwano .dbf lub .xlsx"
        )


def parse_filename(filename: str) -> dict:
    """
    Parsuje nazwe pliku MRPiPS-01 (DBF lub XLSX).

    DBF:  MP1260114  -> rok=2026, miesiac=01, woj=14
    XLSX: 01_2026    -> rok=2026, miesiac=01
    """
    name = os.path.splitext(os.path.basename(filename))[0]
    ext  = os.path.splitext(filename)[1].lower()

    if ext in ('.xlsx', '.xlsm'):
        parts = name.split('_')
        if len(parts) == 2:
            return {
                'raw': name, 'format': 'xlsx',
                'rok': parts[1], 'miesiac': parts[0].zfill(2),
            }

    uname = name.upper()
    if uname.startswith('MP1') and len(uname) >= 9:
        return {
            'raw': name, 'format': 'dbf',
            'rok': '20' + uname[3:5],
            'miesiac': uname[5:7],
            'wojewodztwo': uname[7:9],
        }

    return {'raw': name, 'blad': 'Nieznany format nazwy'}


# ======================================================================
# GLOWNA KLASA
# ======================================================================

class MRPiPS01:
    """
    Klasa do analizy danych MRPiPS-01.
    Obsluguje oba formaty automatycznie:

        mrpips = MRPiPS01('MP1260114.dbf')    # stary format
        mrpips = MRPiPS01('01_2026.xlsx')     # nowy format od 2026
    """

    def __init__(self, filename: str, encoding: str = 'cp1250'):
        self.filename   = filename
        self.format     = os.path.splitext(filename)[1].lower().lstrip('.')
        self.records    = wczytaj_mrpips(filename, encoding)
        self.info_pliku = parse_filename(filename)

        # Indeks: (WGM, DZIAL, TABELA, NRW) -> rekord
        self._index = {}
        for r in self.records:
            key = (r['WGM'], r['DZIAL'], r['TABELA'], r['NRW'])
            self._index[key] = r

        self.powiaty = sorted(set(r['WGM'] for r in self.records if r['DZIAL'] != 'C'))
        sample = next(r for r in self.records if r['DZIAL'] != 'C')
        self.miesiac = sample['M_C']
        self.rok     = sample['ROK']

    def __repr__(self):
        return (f"MRPiPS01(format={self.format.upper()}, rok={self.rok}, "
                f"miesiac={self.miesiac}, powiaty={len(self.powiaty)}, "
                f"rekordy={len(self.records)})")

    # ── Podstawowe pobieranie ────────────────────────────────────────────────

    def get_value(self, powiat: str, dzial: str, tabela: str,
                  nrw: str, kolumna: str) -> int:
        """
        Pobiera wartosc pojedynczej komorki sprawozdania.

        Przyklad:
            mrpips.get_value('1465', '1', '1', '001', 'R1')
            # -> ile osob zarejestrowalo sie w pow. 1465 w miesiacu
        """
        nrw_norm = 'CZ1' if str(nrw).upper() == 'CZ1' else str(nrw).zfill(3)
        key = (str(powiat), str(dzial).upper(), str(tabela), nrw_norm)
        rec = self._index.get(key)
        return (rec.get(kolumna.upper(), 0) or 0) if rec else 0

    def get_row(self, powiat: str, dzial: str, tabela: str, nrw: str) -> dict:
        """Pobiera caly wiersz (R1-R22)."""
        nrw_norm = 'CZ1' if str(nrw).upper() == 'CZ1' else str(nrw).zfill(3)
        key = (str(powiat), str(dzial).upper(), str(tabela), nrw_norm)
        return self._index.get(key, {})

    # ── Sumowania ────────────────────────────────────────────────────────────

    def sum_powiat_list(self, powiaty: list, dzial: str, tabela: str,
                        nrw: str, kolumna: str) -> int:
        """Suma dla listy powiatow."""
        return sum(self.get_value(p, dzial, tabela, nrw, kolumna) for p in powiaty)

    def sum_wojewodztwo(self, dzial: str, tabela: str, nrw: str, kolumna: str) -> int:
        """Suma dla calego wojewodztwa (wszystkie powiaty)."""
        return self.sum_powiat_list(self.powiaty, dzial, tabela, nrw, kolumna)

    # ── Zapytania wysokiego poziomu ───────────────────────────────────────────

    def rejestracje_w_miesiacu(self, powiat: str) -> dict:
        """
        Zarejestrowani w miesiacu sprawozdawczym.
        D1 T1 NRW=001: R1=razem, R2=kobiety
        """
        r = self.get_row(powiat, '1', '1', '001')
        razem   = r.get('R1', 0) or 0
        kobiety = r.get('R2', 0) or 0
        return {'razem': razem, 'kobiety': kobiety, 'mezczyzni': razem - kobiety}

    def bezrobotni_na_koniec_miesiaca(self, powiat: str) -> dict:
        """
        Stan bezrobocia na koniec miesiaca.
        D1 T1 NRW=001: R5=razem, R6=kobiety, R7=z zasilkiem
        """
        r = self.get_row(powiat, '1', '1', '001')
        razem   = r.get('R5', 0) or 0
        kobiety = r.get('R6', 0) or 0
        zasilek = r.get('R7', 0) or 0
        return {
            'razem': razem, 'kobiety': kobiety,
            'mezczyzni': razem - kobiety,
            'z_prawem_do_zasilku': zasilek,
        }

    def wylaczeni_z_ewidencji(self, powiat: str) -> dict:
        """
        Wylaczeni z ewidencji (wyrejestrowani) w miesiacu.
        D1 T2 NRW=034 R1=razem; NRW=035 R1=podjeli prace
        """
        r34 = self.get_row(powiat, '1', '2', '034')
        r35 = self.get_row(powiat, '1', '2', '035')
        razem         = r34.get('R1', 0) or 0
        kobiety       = r34.get('R2', 0) or 0
        podjeli_prace = r35.get('R1', 0) or 0
        return {'razem': razem, 'kobiety': kobiety, 'podjeli_prace': podjeli_prace}

    def get_czas_wypelnienia(self, powiat: str) -> dict:
        """
        Czas wypelnienia formularza (DZIAL C, NRW=CZ1).
        R1=przygotowanie danych (min), R2=wypelnienie (min)
        """
        key = (str(powiat), 'C', '1', 'CZ1')
        rec = self._index.get(key, {})
        return {
            'przygotowanie_min': rec.get('R1', 0) or 0,
            'wypelnienie_min':   rec.get('R2', 0) or 0,
        }

    def sprawozdanie_powiatu(self, powiat: str) -> dict:
        """
        Pelne sprawozdanie dla powiatu jako slownik:
        { dzial: { tabela: { nrw: { R1: val, ..., R22: val } } } }
        """
        result = defaultdict(lambda: defaultdict(dict))
        for rec in self.records:
            if rec['WGM'] == str(powiat):
                d, t, nrw = rec['DZIAL'], rec['TABELA'], rec['NRW']
                result[d][t][nrw] = {f'R{i}': rec.get(f'R{i}', 0) for i in range(1, 23)}
        return dict(result)

    def opis_wiersza(self, dzial: str, tabela: str, nrw: str) -> str:
        """Opis tekstowy wiersza formularza."""
        return OPIS_WIERSZY.get((str(dzial), str(tabela)), {}).get(
            str(nrw).zfill(3), f'Wiersz NRW={nrw}')

    # ── Wydruki ──────────────────────────────────────────────────────────────

    def drukuj_info_powiatu(self, powiat: str):
        """Podsumowanie najwazniejszych danych dla powiatu."""
        reg  = self.rejestracje_w_miesiacu(powiat)
        bezr = self.bezrobotni_na_koniec_miesiaca(powiat)
        wyl  = self.wylaczeni_z_ewidencji(powiat)

        print(f"\n{'='*60}")
        print(f"POWIAT: {powiat}  |  {self.rok}-{self.miesiac}  [{self.format.upper()}]")
        print(f"{'='*60}")
        print(f"Zarejestrowani w mies. : {reg['razem']:6}  (kobiety: {reg['kobiety']:6})")
        print(f"Wylaczeni z ewidencji  : {wyl['razem']:6}  (kobiety: {wyl['kobiety']:6})")
        print(f"  w tym podjeli prace  : {wyl['podjeli_prace']:6}")
        print(f"Stan na koniec mies.   : {bezr['razem']:6}  (kobiety: {bezr['kobiety']:6})")
        print(f"  z prawem do zasilku  : {bezr['z_prawem_do_zasilku']:6}")
        print()
        print("Wybrane kategorie (stan na koniec mies.):")
        for nrw, label in [
            ('005', 'Zamieszkali na wsi'),
            ('017', 'Z pierwszenstwem do form pomocy'),
            ('018', '  - do 30 r.z.'),
            ('020', '  - powyzej 50 r.z.'),
            ('021', '  - dlugotrwale bezrobotni'),
            ('023', '  - niepelnosprawni'),
        ]:
            val = self.get_value(powiat, '1', '1', nrw, 'R5')
            print(f"  {label:<42}: {val:6}")

    def drukuj_zestawienie(self, top_n: int = None, sort_by: str = 'powiat'):
        """
        Zestawienie dla wszystkich powiatow.

        Args:
            top_n:   pokazuje tylko N powiatow (posortowanych wg stanu malejaco)
            sort_by: 'powiat' | 'stan' | 'zarej'
        """
        dane = []
        for p in self.powiaty:
            reg  = self.rejestracje_w_miesiacu(p)
            bezr = self.bezrobotni_na_koniec_miesiaca(p)
            dane.append({
                'powiat': p, 'zarej': reg['razem'], 'zarej_k': reg['kobiety'],
                'stan': bezr['razem'], 'stan_k': bezr['kobiety'],
                'zasilek': bezr['z_prawem_do_zasilku'],
            })

        if sort_by == 'stan':
            dane.sort(key=lambda x: x['stan'], reverse=True)
        elif sort_by == 'zarej':
            dane.sort(key=lambda x: x['zarej'], reverse=True)

        if top_n:
            dane = dane[:top_n]

        print(f"\n{'POWIAT':6}  {'ZAREJ.':>7}  {'K':>6}  "
              f"{'STAN':>8}  {'K':>7}  {'ZASILEK':>8}")
        print(f"{'─'*6}  {'─'*7}  {'─'*6}  {'─'*8}  {'─'*7}  {'─'*8}")
        tot_z = tot_s = 0
        for d in dane:
            print(f"{d['powiat']:6}  {d['zarej']:7}  {d['zarej_k']:6}  "
                  f"{d['stan']:8}  {d['stan_k']:7}  {d['zasilek']:8}")
            tot_z += d['zarej']
            tot_s += d['stan']
        print(f"{'─'*6}  {'─'*7}  {'─'*6}  {'─'*8}  {'─'*7}  {'─'*8}")
        print(f"{'SUMA':6}  {tot_z:7}  {'':6}  {tot_s:8}")

    def drukuj_czas_wypelnienia(self):
        """Zestawienie czasu wypelnienia formularza (DZIAL C) dla wszystkich powiatow."""
        print(f"\n{'='*55}")
        print("CZAS WYPELNIENIA FORMULARZA (DZIAL C, NRW=CZ1)")
        print(f"{'='*55}")
        print(f"{'POWIAT':6}  {'PRZYGOTOWANIE':>15}  {'WYPELNIENIE':>12}")
        print(f"{'─'*6}  {'─'*15}  {'─'*12}")
        for p in self.powiaty:
            c = self.get_czas_wypelnienia(p)
            if c['przygotowanie_min'] or c['wypelnienie_min']:
                print(f"{p:6}  {c['przygotowanie_min']:>12} min  "
                      f"{c['wypelnienie_min']:>9} min")


# ======================================================================
# PRZYKLAD UZYCIA
# ======================================================================

if __name__ == '__main__':

    print("=" * 65)
    print("MRPiPS-01 Parser - demo obu formatow")
    print("=" * 65)

    PLIK_DBF  = 'MP1260114.dbf'
    PLIK_XLSX = '01_2026.xlsx'

    dostepne = [f for f in [PLIK_DBF, PLIK_XLSX] if os.path.exists(f)]
    if not dostepne:
        print("Nie znaleziono plikow. Uruchom z katalogu zawierajacego pliki.")
        exit(1)

    print(f"Znalezione pliki: {dostepne}\n")

    obiekty = {}
    for f in dostepne:
        m = MRPiPS01(f)
        obiekty[f] = m
        print(f"Wczytano: {m}")

    mrpips = list(obiekty.values())[0]

    # -- Podstawowe zapytanie
    print("\n" + "─" * 65)
    print("Zarejestrowani w miesiacu w pow. 1465:")
    val = mrpips.get_value('1465', '1', '1', '001', 'R1')
    print(f"  D1 T1 NRW=001 R1 = {val} osob")

    # -- Info o powiatach
    mrpips.drukuj_info_powiatu('1401')
    mrpips.drukuj_info_powiatu('1465')

    # -- Sumy wojewodzkie (weryfikacja z tablicami kontrolnymi)
    print("─" * 65)
    print("Sumy dla wojewodztwa:")
    for nrw, label in [
        ('001', 'Ogolem zarej. w mies.'),
        ('005', 'Zamieszkali na wsi'),
        ('021', 'Dlugotrwale bezrobotni'),
        ('023', 'Niepelnosprawni'),
    ]:
        print(f"  NRW={nrw} ({label}): {mrpips.sum_wojewodztwo('1', '1', nrw, 'R1')}")

    # -- Zestawienie top-10
    print("\nTop 10 powiatow wg stanu bezrobotnych:")
    mrpips.drukuj_zestawienie(top_n=10, sort_by='stan')

    # -- Czas wypelnienia
    mrpips.drukuj_czas_wypelnienia()

    # -- Nazwy plikow
    print("\n" + "─" * 65)
    print("Parsowanie nazw plikow:")
    for f in dostepne:
        print(f"  {os.path.basename(f):20} -> {parse_filename(f)}")

    # -- Porownanie obu formatow jesli oba dostepne
    if len(obiekty) == 2:
        print("\n" + "─" * 65)
        print("Porownanie DBF vs XLSX (te same wartosci?):")
        m_dbf  = obiekty[PLIK_DBF]
        m_xlsx = obiekty[PLIK_XLSX]
        for powiat in ['1401', '1425', '1465']:
            v_dbf  = m_dbf.get_value(powiat, '1', '1', '001', 'R1')
            v_xlsx = m_xlsx.get_value(powiat, '1', '1', '001', 'R1')
            status = "OK" if v_dbf == v_xlsx else "ROZNICA!"
            print(f"  Powiat {powiat}: DBF={v_dbf:6}  XLSX={v_xlsx:6}  {status}")
